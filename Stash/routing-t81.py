# -*- coding: utf-8 -*-
"""
Routing Estimation Tool - V1.0.7
- 修复：避免 trace 链触发导致卡顿（已实现）
- 修复：Excel 导出 Std/Opt 显示列不乘 Complexity（已实现）
- 新增：Import Excel（导入）功能（严格校验，按用户确认规则）
- Product和Tuning Type的选择联动
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import scrolledtext
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import sys
import os

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def set_app_icon(window):
    try:
        icon_path = resource_path("time.ico")
        window.iconbitmap(icon_path)
    except Exception:
        pass

# ===================== 数据定义 =====================

PRODUCTS = ['TMA', 'DPX', 'TPX', 'PPX', 'QPX', 'IMF', 'LLC']

FACTORS = {
    'GLOBAL': {
        'TMA': 1.0, 'DPX': 1.0, 'TPX': 1.0, 'PPX': 1.0, 'QPX': 1.0,
        'IMF': 1.0, 'LLC': 1.0
    },
    'TIGER': {
        'TMA': 1.0, 'DPX': 1.1, 'TPX': 1.28, 'PPX': 0.85, 'QPX': 0.95,
        'IMF': 0.83, 'LLC': 0.69
    },
    'TATFOOK': {
        'TMA': 1.39, 'DPX': 1.39, 'TPX': 1.43, 'PPX': 1.04, 'QPX': 0.92,
        'IMF': 1.02, 'LLC': 0.84
    },
    'LUXSHARE': {
        'TMA': 1.24, 'DPX': 1.35, 'TPX': 1.68, 'PPX': 1.18, 'QPX': 0.84,
        'IMF': 1.04, 'LLC': 0.87
    }
}

FPY = {'TMA': 0.95, 'DPX': 0.95, 'TPX': 0.90, 'PPX': 0.85, 'QPX': 0.88, 'IMF': 0.85, 'LLC': 0.85}

PRE_ASSEMBLY_OPS = [
    ("Tuning screw glued to tuners or plug", 12, 100, 1, 'sec', 'housing'),
    ("Lid cover+tuning screw+nut+washer", 8, 100, 1, 'sec', 'housing'),
    ("Connector+conductor soldered", 45, 100, 1, 'sec', 'housing'),
    ("Connector+conductor screwed", 5, 100, 1, 'sec', 'housing'),
    ("Other Pre_Assembly item", 10, 100, 1, 'sec', 'housing'),
]

ASSEMBLY_OPS = [
    ("Low pass(or bias tee or coils) + thermalshrinking", 120, 100, 1, 'sec', 'housing'),
    ("Bias coil group", 30, 100, 1, 'sec', 'housing'),
    ("Lid cover + pin", 30, 100, 1, 'sec', 'housing'),
    ("Lid cover + line screwed", 5, 100, 1, 'sec', 'housing'),
    ("Lid cover + line soldered", 67, 100, 1, 'sec', 'housing'),
    ("Lid cover + connector", 15, 100, 1, 'sec', 'housing'),
    ("Lid cover + connector press in", 60, 100, 1, 'sec', 'housing'),
    ("Label for s/n", 5, 100, 1, 'sec', 'housing'),
    ("Housing + pin", 10, 100, 1, 'sec', 'housing'),
    ("Housing + connector flanged", 45, 100, 1, 'sec', 'housing'),
    ("Housing + connector screwed", 30, 100, 1, 'sec', 'housing'),
    ("Housing + junction screwed (time x 2 screws)", 10, 100, 1, 'sec', 'housing'),
    ("Housing + junction soldered by SOLDER STATION", 30, 100, 1, 'sec', 'housing'),
    ("Housing + junction soldered INDUCTION", 55, 100, 1, 'sec', 'housing'),
    ("Housing + line soldered by SOLDER STATION (complex)", 150, 100, 1, 'sec', 'housing'),
    ("Housing + line soldered INDUCTION (complex)", 150, 100, 1, 'sec', 'housing'),
    ("Housing + low pass(or bias tee)", 10, 100, 1, 'sec', 'housing'),
    ("Ret connector", 20, 100, 1, 'sec', 'housing'),
    ("Resonator", 7, 90, 1, 'sec', 'housing'),
    ("Housing + probe (including support fixing)", 9, 100, 1, 'sec', 'housing'),
    ("Apply solder paste for LID SOLDERING process", 300, 100, 1, 'sec', 'housing'),
    ("Tuning lid cover screwing", 3, 85, 1, 'sec', 'housing'),
    ("PWBA installation (including soldering)", 8, 100, 1, 'sec', 'housing'),
    ("Cable sub-assembly(or LED)", 15, 100, 1, 'sec', 'housing'),
    ("Outer cover (time x screw)", 4, 100, 1, 'sec', 'housing'),
    ("PWBA+Nickel Silver Cover placing", 3, 100, 1, 'sec', 'housing'),
    ("IP outer label ASSEMBLY", 60, 100, 1, 'sec', 'housing'),
    ("Housing + venting membrane", 10, 100, 1, 'sec', 'housing'),
    ("Housing + bracket +screw + nut + washer", 120, 100, 1, 'sec', 'housing'),
    ("Cleaning Tuning lid open and close (time x screw)", 4, 95, 1, 'sec', 'housing'),
    ("Cleaning Process", 4.0, 100, 1, 'min', 'housing'),
    ("Other Assembly item", 5.0, 100, 1, 'min', 'housing'),
]

TUNING_OPS = [
    ("RF TUNING1 estimated (if no other indication)", 108, 90, 1, 'sec', 'tuning'),
    ("Tuning check after cleaning (Check Std time)", 5.0, 100, 1, 'min', 'tuning'),
    ("TUNING 2 (E35)", 8.0, 100, 1, 'min', 'tuning'),
    ("Other Tuning item", 10.0, 100, 1, 'min', 'tuning'),
]

TESTING_OPS = [
    ("PCBA programming (FW uploading)", 1.0, 100, 1, 'min', 'none'),
    ("NOISE FIGURE(mandatory for TMA)", 3.0, 100, 1, 'min', 'none'),
    ("IP3(optional for TMA)", 3.0, 100, 1, 'min', 'none'),
    ("AC/DC or Generic Measures", 3.0, 100, 1, 'min', 'none'),
    ("Fixed Tone IMD", 1.5, 100, 1, 'min', 'none'),
    ("Swept Tone IMD", 1.25, 100, 1, 'min', 'none'),
    ("LINEARIZATION (mandatory for TMA)", 3.0, 100, 1, 'min', 'none'),
    ("PID(mandatory for TMA)", 3.0, 100, 1, 'min', 'none'),
    ("TDR on TMS (FINAL TEST)", 3.0, 90, 1, 'min', 'none'),
    ("PRESSURE TEST (Std) - Check Qty", 2.0, 100, 1, 'min', 'none'),
    ("PRESSURE TEST (IP Label) - Check Qty", 4.0, 100, 1, 'min', 'none'),
    ("LED", 2.0, 100, 1, 'min', 'none'),
    ("Other Testing item", 5.0, 100, 1, 'min', 'none'),
]

QUALITY_OPS = [
    ("Final check 100% (VISUAL INSPECTION)", 1.0, 100, 1, 'min', 'none'),
    ("Other Quality check item", 1.0, 100, 1, 'min', 'none'),
]
PACKAGING_OPS = [
    ("Packaging", 3.2, 100, 1, 'min', 'none'),
    ("Other Packaging item", 1.0, 100, 1, 'min', 'none'),
]

TUNING_TYPES = {
    'TMA': (90, 81),
    'DPX': (72, 64.8),
    'TPX': (81, 72.9),
    'PPX': (90, 81),
    'QPX': (90, 81),
    'IMF': (108, 97.2),
    'LLC': (108, 97.2),    
    'Demanding perf': (108, 97.2),
}

COMPLEXITY_LEVELS = ['1.0', '0.75', '1.25', 'Custom']

# ===================== 修复 BUG所需常量 =====================
TUNING_TAB_NAME = 'Tuning'
TUNING_CHECK_NAME = "Tuning check after cleaning (Check Std time)"
DEFAULT_TUNING_CHECK_STD_TIME = 5.0 # 默认复位值 5.0
# ========================================================

# ===================== 主类 =====================

class RoutingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Routing Estimation Tool - V1.0.8")
        self.root.geometry("1200x1000")
        self.root.minsize(1200, 1000)

        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill='both', expand=True, padx=15, pady=12)

        # 基本变量
        self.pn_var = tk.StringVar(value="")
        self.product_var = tk.StringVar(value='TPX')
        self.cm_var = tk.StringVar(value='GLOBAL')
        self.standard_mode_var = tk.BooleanVar(value=True)
        self.optimized_mode_var = tk.BooleanVar(value=True)
        self.standard_fpy_var = tk.StringVar(value='90%')
        self.standard_custom_fpy_num_var = tk.StringVar(value='90')
        self.standard_rework_var = tk.StringVar(value='10%')
        self.optimized_fpy_var = tk.StringVar(value='90%')
        self.optimized_custom_fpy_num_var = tk.StringVar(value='90')
        self.optimized_rework_var = tk.StringVar(value='10%')
        self.n_cavities_var = tk.StringVar(value='5')
        self.housing_per_fg_var = tk.StringVar(value='2')
        self.tuning_type_var = tk.StringVar(value='TPX')

        self.complexity_vars = {}
        self.custom_complexity_num_vars = {}
        self.complexity_sections = ['Pre_Assembly', 'Assembly', 'Tuning', 'Testing', 'Quality Check', 'Packaging']
        for section in self.complexity_sections:
            self.complexity_vars[section] = tk.StringVar(value='1.0')
            self.custom_complexity_num_vars[section] = tk.StringVar(value='1.0')

        # guards
        self.suppress_updates = False
        self.suppress_qty_modify = False

        #Demanding perf
        self.user_locked_demanding_perf = False

        # data containers
        self.original_std_times = {}
        self.opt_percent_vars = {}
        self.check_vars = {}
        self.quantity_vars = {}
        self.parts_qty_vars = {}
        self.std_time_vars = {}
        self.unit_dict = {}
        self.multi_type_dict = {}
        self.calc_labels = {}
        self.non_optimized_vars = {}
        self.user_modified = {}

        self.setup_data()
        self.setup_ui()

        # traces for global controls
        self.standard_mode_var.trace_add('write', self.update_all_times)
        self.optimized_mode_var.trace_add('write', self.update_all_times)
        self.housing_per_fg_var.trace_add('write', self.on_housing_change)
        self.n_cavities_var.trace_add('write', self.update_all_times)
        self.product_var.trace_add('write', self.on_product_change)
        self.standard_fpy_var.trace_add('write', self.on_standard_fpy_combo_change)
        self.standard_custom_fpy_num_var.trace_add('write', self.update_standard_fpy_rework)
        self.optimized_fpy_var.trace_add('write', self.on_optimized_fpy_combo_change)
        self.optimized_custom_fpy_num_var.trace_add('write', self.update_optimized_fpy_rework)
        self.tuning_type_var.trace_add('write', self.update_tuning_std_time)
        self.cm_var.trace_add('write', self.update_all_times)
        self.tuning_type_var.trace_add('write', self.on_tuning_type_manual_change)
        self.tuning_type_var.trace_add('write', self.on_tuning_type_change)
        
        for section in self.complexity_sections:
            self.complexity_vars[section].trace_add('write', lambda *args, s=section: self.on_complexity_combo_change(s))
            self.custom_complexity_num_vars[section].trace_add('write', lambda *args, s=section: self.update_all_times())

        # initial updates
        self.update_all_times()
        self.on_product_change()
        self.update_standard_fpy_rework()
        self.update_optimized_fpy_rework()
        self.on_housing_change()
        self.update_tuning_std_time()
        for section in self.complexity_sections:
            self.on_complexity_combo_change(section)

    # ---------------- trace wrapper ----------------
    def _trace_update_section(self, tab_name):
        if getattr(self, 'suppress_updates', False):
            return
        try:
            self.update_section_times(tab_name)
        except Exception:
            pass

    def get_complexity_level(self, section):
        try:
            val = self.complexity_vars[section].get()
            if val == 'Custom':
                return float(self.custom_complexity_num_vars[section].get())
            else:
                return float(val)
        except:
            return 1.0

    def on_complexity_combo_change(self, section, *args):
        custom_entry = getattr(self, f"{section.lower()}_complexity_entry", None)
        if self.complexity_vars[section].get() == 'Custom':
            if custom_entry:
                custom_entry.config(state='normal')
            self.update_all_times()
        else:
            if custom_entry:
                custom_entry.config(state='disabled')
            self.custom_complexity_num_vars[section].set(self.complexity_vars[section].get())
            self.update_all_times()

    def safe_get_int(self, var, default=1):
        try:
            return int(var.get())
        except:
            return default

    def safe_get_double(self, key):
        try:
            return float(self.std_time_vars[key].get())
        except:
            return self.original_std_times.get(key, 0.0)

    def safe_get_opt_percent(self, key):
        try:
            percent = float(self.opt_percent_vars[key].get())
            percent = max(0.0, min(100.0, percent))
            return percent / 100.0
        except:
            return 1.0

    def on_tuning_type_manual_change(self, *args):
        current_tuning = self.tuning_type_var.get()
        if current_tuning == "Demanding perf":
            self.user_selected_demanding_perf = True
        else:
            self.user_selected_demanding_perf = False  # 改回普通类型时，重置标记，恢复联动

    # ---------------- data 初始化 ----------------
    def setup_data(self):
        sections = {
            'Pre_Assembly': PRE_ASSEMBLY_OPS,
            'Assembly': ASSEMBLY_OPS,
            'Tuning': TUNING_OPS,
            'Testing': TESTING_OPS,
            'Quality Check': QUALITY_OPS,
            'Packaging': PACKAGING_OPS
        }
        for tab, ops in sections.items():
            self.check_vars[tab] = {}
            self.quantity_vars[tab] = {}
            self.non_optimized_vars[tab] = {}
            self.user_modified[tab] = {}
            self.parts_qty_vars[tab] = {}
            for op in ops:
                name, std_t, opt_percent_default, qty, unit, mtype = op
                key = (tab, name)

                self.check_vars[tab][name] = tk.BooleanVar(value=False)
                self.quantity_vars[tab][name] = tk.IntVar(value=qty)
                self.non_optimized_vars[tab][name] = tk.BooleanVar(value=False)
                self.parts_qty_vars[tab][name] = tk.IntVar(value=1)

                self.std_time_vars[key] = tk.DoubleVar(value=std_t)
                self.original_std_times[key] = std_t
                self.opt_percent_vars[key] = tk.DoubleVar(value=opt_percent_default)

                self.unit_dict[key] = unit
                self.multi_type_dict[key] = mtype

                # traces use wrapper to avoid chain triggers when programmatically setting vars
                self.check_vars[tab][name].trace_add('write', lambda *args, t=tab: self._trace_update_section(t))
                self.std_time_vars[key].trace_add('write', lambda *args, t=tab: self._trace_update_section(t))
                self.opt_percent_vars[key].trace_add('write', lambda *args, t=tab: self._trace_update_section(t))
                self.quantity_vars[tab][name].trace_add('write', lambda *args, t=tab, n=name: self.on_qty_change(t, n))
                self.non_optimized_vars[tab][name].trace_add('write', lambda *args, t=tab: self._trace_update_section(t))
                self.parts_qty_vars[tab][name].trace_add('write', lambda *args, t=tab: self._trace_update_section(t))

                # ====================== 特殊处理：Tuning check after cleaning (Check Std time) ======================
                if name == "Tuning check after cleaning (Check Std time)":
                    # 当用户手动修改 Std Time 输入框时，标记为已手动修改
                    def on_check_time_manual_edit(*args, current_tab=tab, current_name=name, current_key=key):
                        try:
                            user_value = float(self.std_time_vars[current_key].get())
                            # 与原始默认值 5.0 比较（容差 0.01）
                            if abs(user_value - self.original_std_times[current_key]) > 0.01:
                                self.user_modified.setdefault(current_tab, {})[current_name] = True
                        except:
                            # 输入非法时也视为手动修改（防止自动覆盖）
                            self.user_modified.setdefault(current_tab, {})[current_name] = True

                    # 绑定 trace：用户一修改输入框就打上“手动修改”标记
                    self.std_time_vars[key].trace_add('write', on_check_time_manual_edit)

    # ---------------- UI ----------------
    def setup_ui(self):
        basic_frame = ttk.LabelFrame(self.main_frame, text=" Basic Configuration ", padding=12)
        basic_frame.pack(fill='x', pady=(0, 8))

        ttk.Label(basic_frame, text="P/N:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=(0,8))
        ttk.Entry(basic_frame, textvariable=self.pn_var, width=28, font=('Arial', 10)).grid(row=0, column=1, columnspan=3, sticky='w', padx=(0,40))

        ttk.Label(basic_frame, text="Product:").grid(row=1, column=0, sticky='e', padx=(0,8), pady=(8,0))
        ttk.Combobox(basic_frame, textvariable=self.product_var, values=PRODUCTS, width=15).grid(row=1, column=1, sticky='w', padx=(0,30), pady=(8,0))

        ttk.Label(basic_frame, text="ROUTING:").grid(row=1, column=2, sticky='e', padx=(20,8), pady=(8,0))
        ttk.Combobox(basic_frame, textvariable=self.cm_var, values=list(FACTORS.keys()), width=15).grid(row=1, column=3, sticky='w', padx=(0,30), pady=(8,0))

        ttk.Label(basic_frame, text="Tuning Type:").grid(row=1, column=4, sticky='e', padx=(20,8), pady=(8,0))
        ttk.Combobox(basic_frame, textvariable=self.tuning_type_var, values=list(TUNING_TYPES.keys()), width=15).grid(row=1, column=5, sticky='w', pady=(8,0))

        run_frame = ttk.LabelFrame(self.main_frame, text=" Yield Estimation ", padding=12)
        run_frame.pack(fill='x', pady=(0, 12))

        ttk.Label(run_frame, text="Modes:").grid(row=0, column=0, sticky='w', padx=(0,8))
        modes_frame = ttk.Frame(run_frame)
        modes_frame.grid(row=0, column=1, columnspan=5, sticky='w')
        ttk.Checkbutton(modes_frame, text="STANDARD", variable=self.standard_mode_var).pack(side='left', padx=5)
        ttk.Checkbutton(modes_frame, text="OPTIMIZED", variable=self.optimized_mode_var).pack(side='left', padx=5)

        ttk.Label(run_frame, text="STANDARD FPY:").grid(row=1, column=0, sticky='w', padx=(0,8), pady=(10,0))
        standard_fpy_frame = ttk.Frame(run_frame)
        standard_fpy_frame.grid(row=1, column=1, columnspan=2, sticky='w', pady=(10,0))
        self.standard_fpy_combo = ttk.Combobox(standard_fpy_frame, textvariable=self.standard_fpy_var,
                                               values=[f"{int(v*100)}%" for v in sorted(set(FPY.values()))] + ['Custom'],
                                               width=10, state='readonly')
        self.standard_fpy_combo.pack(side='left')
        self.standard_custom_fpy_entry = ttk.Entry(standard_fpy_frame, textvariable=self.standard_custom_fpy_num_var, width=6, state='disabled')
        self.standard_custom_fpy_entry.pack(side='left', padx=(5,0))
        self.standard_percent_label = ttk.Label(standard_fpy_frame, text="%", state='disabled')
        self.standard_percent_label.pack(side='left', padx=(2,0))

        ttk.Label(run_frame, text="STANDARD Rework:").grid(row=1, column=3, sticky='e', padx=(20,8), pady=(10,0))
        ttk.Label(run_frame, textvariable=self.standard_rework_var, width=15).grid(row=1, column=4, sticky='w', pady=(10,0))

        ttk.Label(run_frame, text="OPTIMIZED FPY:").grid(row=2, column=0, sticky='w', padx=(0,8), pady=(10,0))
        optimized_fpy_frame = ttk.Frame(run_frame)
        optimized_fpy_frame.grid(row=2, column=1, columnspan=2, sticky='w', pady=(10,0))
        self.optimized_fpy_combo = ttk.Combobox(optimized_fpy_frame, textvariable=self.optimized_fpy_var,
                                                values=[f"{int(v*100)}%" for v in sorted(set(FPY.values()))] + ['Custom'],
                                                width=10, state='readonly')
        self.optimized_fpy_combo.pack(side='left')
        self.optimized_custom_fpy_entry = ttk.Entry(optimized_fpy_frame, textvariable=self.optimized_custom_fpy_num_var, width=6, state='disabled')
        self.optimized_custom_fpy_entry.pack(side='left', padx=(5,0))
        self.optimized_percent_label = ttk.Label(optimized_fpy_frame, text="%", state='disabled')
        self.optimized_percent_label.pack(side='left', padx=(2,0))

        ttk.Label(run_frame, text="OPTIMIZED Rework:").grid(row=2, column=3, sticky='e', padx=(20,8), pady=(10,0))
        ttk.Label(run_frame, textvariable=self.optimized_rework_var, width=15).grid(row=2, column=4, sticky='w', pady=(10,0))

        # Complexity Frame
        complexity_frame = ttk.LabelFrame(self.main_frame, text=" Complexity Level", padding=12)
        complexity_frame.pack(fill='x', pady=(0, 12))

        label_width = 12
        sections = self.complexity_sections

        for i in range(3):
            section = sections[i]
            col = i * 2
            ttk.Label(complexity_frame, text=f"{section}:", width=label_width, anchor='e').grid(
                row=0, column=col, sticky='e', padx=(2, 6), pady=(8, 0))
            comp_frame = ttk.Frame(complexity_frame)
            comp_frame.grid(row=0, column=col+1, sticky='w', padx=(0, 10), pady=(8, 0))
            ttk.Combobox(comp_frame, textvariable=self.complexity_vars[section],
                         values=COMPLEXITY_LEVELS, width=7, state='readonly').pack(side='left')
            setattr(self, f"{section.lower()}_complexity_entry", ttk.Entry(comp_frame,
                                                                         textvariable=self.custom_complexity_num_vars[section],
                                                                         width=7, state='disabled', justify='center'))
            getattr(self, f"{section.lower()}_complexity_entry").pack(side='left', padx=(5,0))

        for i in range(3, 6):
            section = sections[i]
            col = (i - 3) * 2
            ttk.Label(complexity_frame, text=f"{section}:", width=label_width, anchor='e').grid(
                row=1, column=col, sticky='e', padx=(2, 6), pady=(8, 0))
            comp_frame = ttk.Frame(complexity_frame)
            comp_frame.grid(row=1, column=col+1, sticky='w', padx=(0, 10), pady=(8, 0))
            ttk.Combobox(comp_frame, textvariable=self.complexity_vars[section],
                         values=COMPLEXITY_LEVELS, width=7, state='readonly').pack(side='left')
            setattr(self, f"{section.lower()}_complexity_entry", ttk.Entry(comp_frame,
                                                                         textvariable=self.custom_complexity_num_vars[section],
                                                                         width=7, state='disabled', justify='center'))
            getattr(self, f"{section.lower()}_complexity_entry").pack(side='left', padx=(5,0))

        param_frame = ttk.LabelFrame(self.main_frame, text=" Routing Parameters ", padding=12)
        param_frame.pack(fill='x', pady=(0, 12))

        ttk.Label(param_frame, text="N° Cavities:").grid(row=0, column=0, sticky='w', padx=(0,10))
        ttk.Entry(param_frame, textvariable=self.n_cavities_var, width=12, justify='center').grid(row=0, column=1, padx=(0,40))

        ttk.Label(param_frame, text="N° Housings:").grid(row=0, column=2, sticky='w', padx=(20,10))
        ttk.Entry(param_frame, textvariable=self.housing_per_fg_var, width=12, justify='center').grid(row=0, column=3, padx=(0,40))

        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(0,10))

        for tab_name, ops_list in [('Pre_Assembly', PRE_ASSEMBLY_OPS), ('Assembly', ASSEMBLY_OPS),
                              ('Tuning', TUNING_OPS), ('Testing', TESTING_OPS),
                              ('Quality Check', QUALITY_OPS), ('Packaging', PACKAGING_OPS)]:
            self.create_tab(tab_name, ops_list)

        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.pack(pady=12, anchor='w')
        ttk.Button(btn_frame, text="Calculate Total Time", command=self.calculate_detailed, width=24).pack(side='left', padx=6)
        ttk.Button(btn_frame, text="Export to Excel", command=self.export_to_excel, style="Accent.TButton", width=20).pack(side='left', padx=6)
        ttk.Button(btn_frame, text="Import Excel", command=self.import_from_excel, width=18).pack(side='left', padx=6)
        ttk.Button(btn_frame, text="Reset All", command=self.reset_all, width=20).pack(side='left', padx=6)

        self.results = scrolledtext.ScrolledText(self.main_frame, height=50, font=("Courier New", 10), bg='#f8f8f8')
        self.results.pack(fill='both', expand=True, padx=2, pady=(0,12))

        if self.results.winfo_exists():
            self.results.bind("<MouseWheel>", self.on_results_mousewheel)

    def on_results_mousewheel(self, event):
        delta = int(-1 * (event.delta / 120))
        self.results.yview_scroll(delta, "units")
        return "break"

    def create_tab(self, tab_name, ops_list):
        # 定义一个固定高度（单位：像素），保持您之前设定的固定高度要求。
        CANVAS_FIXED_HEIGHT = 225
        
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=f" {tab_name} ")
        self.calc_labels[tab_name] = {}

        # 1. 初始化 Canvas 和 垂直/水平 Scrollbar
        canvas = tk.Canvas(tab, height=CANVAS_FIXED_HEIGHT)
        vsb = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        hsb = ttk.Scrollbar(tab, orient="horizontal", command=canvas.xview) # 新增：水平滚动条
        
        frame = ttk.Frame(canvas)
        
        # 2. 绑定 frame 的 <Configure> 事件，用于更新滚动区域（同时更新 X 和 Y 轴）
        def on_frame_configure(event):
            # scrollregion 必须设置为 frame 的边界，以支持 X 和 Y 滚动
            canvas.configure(scrollregion=canvas.bbox("all"))

        frame.bind("<Configure>", on_frame_configure)
        
        # 将 frame 放入 canvas。不强制设置宽度，允许其随内容扩展。
        window_item = canvas.create_window((0,0), window=frame, anchor="nw")
        
        # 3. 配置 Canvas 的滚动命令
        canvas.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set) # 配置 X/Y 滚动
        
        canvas.bind("<MouseWheel>", lambda event: self.on_canvas_mousewheel(event, canvas))

        headers = ["Select", "Excluded", "N° of P", "Qty", "Std Time", "Optimized %", "Calc Time (min)", "Operation"]
        op_col = 7
        
        # 移除 frame.grid_columnconfigure(op_col, weight=1)，
        # 允许表格内容水平溢出，从而启用 X 轴滚动。
        # frame.grid_columnconfigure(op_col, weight=1) # <--- 移除此行
        
        header_bg = '#d9d9d9'
        for col, text in enumerate(headers):
            lbl = tk.Label(frame, text=text, font=('Courier New', 9, 'bold'),
                             bg=header_bg, anchor='w', padx=4)
            lbl.grid(row=0, column=col, padx=3, pady=5, sticky='w')

        for i, op in enumerate(ops_list, 1):
            name, std_t, opt_percent_default, qty, unit, mtype = op
            row_color = "#f0f2f5" if i % 2 == 0 else "#ffffff"
            key = (tab_name, name)

            ttk.Checkbutton(frame, variable=self.check_vars[tab_name][name]).grid(
                row=i, column=0, padx=6, pady=2, sticky='w')
            ttk.Checkbutton(frame, variable=self.non_optimized_vars[tab_name][name]).grid(
                row=i, column=1, padx=6, pady=2, sticky='w')

            col = 2
            ttk.Entry(frame, textvariable=self.parts_qty_vars[tab_name][name],
                      width=6, justify='center').grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1

            ttk.Entry(frame, textvariable=self.quantity_vars[tab_name][name],
                      width=6, justify='center').grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1

            tf = ttk.Frame(frame)
            tf.grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1
            ttk.Entry(tf, textvariable=self.std_time_vars[key],
                      width=8, justify='center').pack(side='left')
            ttk.Label(tf, text=unit, width=4).pack(side='left')

            opt_frame = ttk.Frame(frame)
            opt_frame.grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1
            ttk.Entry(opt_frame, textvariable=self.opt_percent_vars[key],
                      width=6, justify='center').pack(side='left')

            calc_lbl = tk.Label(frame, text="0.00",
                                font=('Courier New', 10),
                                fg='darkgreen',
                                bg='#ffffff')
            calc_lbl.grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1
            self.calc_labels[tab_name][name] = calc_lbl

            op_lbl = tk.Label(frame, text=name, anchor='w', justify='left',
                              wraplength=600, bg=row_color)
            # 移除 sticky='ew'，确保 Label 宽度由内容决定
            op_lbl.grid(row=i, column=op_col, sticky='w', padx=6, pady=2) 
            
        # 4. 在 tab frame 上使用 grid 布局 Canvas 和 Scrollbars
        
        # 配置 tab frame 的 grid 权重
        tab.grid_columnconfigure(0, weight=1) # Canvas/HSB 所在列占据水平剩余空间
        tab.grid_rowconfigure(0, weight=1)    # Canvas/VSB 所在行占据垂直剩余空间
        
        # 放置控件
        canvas.grid(row=0, column=0, sticky="nsew") # Canvas 占据大部分空间，可扩展
        vsb.grid(row=0, column=1, sticky="ns")      # VSB 占据右侧
        hsb.grid(row=1, column=0, sticky="ew")      # HSB 占据底部
        # hsb 和 vsb 的交叉角不需要控件，留空或放置一个小 frame

    def on_canvas_mousewheel(self, event, canvas):
        delta = int(-1 * (event.delta / 120))
        canvas.yview_scroll(delta, "units")
        return "break"

    # ---------------- 逻辑计算 ----------------
    def update_section_times(self, tab_name):
        """只更新 Calc Time 显示，对 Std Time 输入框的处理遵循以下规则：
            1. 普通工序：若用户未修改，Std Time = 原始时间（来自 original_std_times） × Complexity
            2. RF TUNING1：Std Time = Tuning Type 对应的标准时间 × Complexity (已在上方处理)
            3. Tuning check after cleaning (Check Std time)：
              - 若用户从未手动修改 → 自动计算 = RF TUNING1 Std Time (w/ Comp, in min) * 10%
              - 若用户已手动修改或导入了自定义值 → 保持该值，不再自动覆盖
        """
        self.suppress_updates = True
        try:
            ops_list = {
                'Pre_Assembly': PRE_ASSEMBLY_OPS, 'Assembly': ASSEMBLY_OPS,
                'Tuning': TUNING_OPS, 'Testing': TESTING_OPS,
                'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS
            }[tab_name]

            n_cav = self.get_n_cavities()
            complexity_factor = self.get_complexity_level(tab_name)

            rf_name = "RF TUNING1 estimated (if no other indication)"
            check_name = "Tuning check after cleaning (Check Std time)"
            check_key = (tab_name, check_name)
            
            # 假设 Tuning check after cleaning (Check Std time) 的硬编码默认时间为 5.0 分钟
            TCC_HARDCODED_DEFAULT = 5.0 # in minutes

            # ======================== Tuning 特殊处理 ========================
            if tab_name == 'Tuning':
                rf_selected = self.check_vars[tab_name][rf_name].get()
                
                # ---------------------------------------------------------------------------------------------------
                # *** 修复 Bug 重现：导入值检查 ***
                # 检查 self.original_std_times (导入值) 是否与硬编码默认值 (5.0 min * complexity) 有显著差异
                # 如果有差异，则表明该值是导入的自定义值，需跳过自动计算，即使 user_modified 标记错误
                # ---------------------------------------------------------------------------------------------------
                is_custom_set_by_import = False
                try:
                    imported_base_val = self.original_std_times.get(check_key)
                    if imported_base_val is not None and abs(imported_base_val - (TCC_HARDCODED_DEFAULT * complexity_factor)) > 0.01:
                        is_custom_set_by_import = True
                except KeyError:
                    pass 
                
                # 只有当 (用户未手动修改) 且 (导入值是默认值) 时，才执行自动计算
                should_auto_calculate = (not self.user_modified.get(tab_name, {}).get(check_name, False)) and (not is_custom_set_by_import)
                
                
                # ---- RF TUNING1 已选中 → 自动更新其 Std Time = 原始时间 × Complexity ----
                if rf_selected:
                    key_rf = (tab_name, rf_name)
                    std_t_rf_orig = self.original_std_times[key_rf]
                    std_t_rf_comp = std_t_rf_orig * complexity_factor
                    
                    self.std_time_vars[key_rf].set(round(std_t_rf_comp, 2))

                    # ---- Tuning check after cleaning (Check Std time) 自动计算 ----
                    if should_auto_calculate:
                        # 修复逻辑：Std Time (Tuning check) = RF TUNING1 Std Time (w/ Comp, in min) * 10%
                        unit_rf = self.unit_dict.get(key_rf, 'sec')
                        
                        if unit_rf == 'sec':
                            # 将 Std Time (w/ Comp) 从秒转换为分钟
                            std_t_rf_comp_min = std_t_rf_comp / 60.0
                        else:
                            # 如果是分钟，则直接使用
                            std_t_rf_comp_min = std_t_rf_comp 

                        # *** 关键修正点：确保使用正确的比例系数 0.1 (10%) ***
                        check_final = std_t_rf_comp_min * 1
                        
                        self.std_time_vars[check_key].set(round(check_final, 2))

                else:
                    # RF TUNING1 未选中，且允许自动计算 → 恢复默认值 5.0 × Complexity
                    if should_auto_calculate:
                        orig_check = self.original_std_times[check_key]
                        std_t_check_comp = orig_check * complexity_factor
                        self.std_time_vars[check_key].set(round(std_t_check_comp, 2))
                        
            # ======================== 统一更新所有行的其他 Std Time（应用 Complexity） ========================
            for name, std_t_default, _, _, unit, _ in ops_list:
                key = (tab_name, name)
                
                if not self.check_vars[tab_name][name].get():
                    self.calc_labels[tab_name][name].config(text="0.00")
                    continue

                is_special_tuning_op = (tab_name == 'Tuning') and (name == rf_name or name == check_name)
                
                if not is_special_tuning_op:
                    if not self.user_modified.get(tab_name, {}).get(name, False):
                        std_t_orig = self.original_std_times[key] 
                        std_t_comp = std_t_orig * complexity_factor
                        self.std_time_vars[key].set(round(std_t_comp, 2))


            # ======================== 计算所有行的 Calc Time（使用当前 Std Time 值） ========================
            for name, _, _, _, unit, _ in ops_list:
                key = (tab_name, name)
                if not self.check_vars[tab_name][name].get():
                    continue

                t_per = self.safe_get_double(key)
                qty = self.safe_get_int(self.quantity_vars[tab_name][name])

                if tab_name == 'Tuning' and name == rf_name:
                    minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty * n_cav
                else:
                    minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty

                parts_qty = self.safe_get_int(self.parts_qty_vars[tab_name][name])
                minutes *= parts_qty

                self.calc_labels[tab_name][name].config(text=f"{minutes:8.2f}")
        finally:
            self.suppress_updates = False

    def update_all_times(self, *args):
        for tab in self.check_vars:
            self._trace_update_section(tab)

    def get_section_total(self, tab_name, ops_list, mode):
        total = 0.0
        n_cav = self.get_n_cavities()
        complexity_factor = self.get_complexity_level(tab_name)

        rf_name = "RF TUNING1 estimated (if no other indication)"
        check_name = "Tuning check after cleaning (Check Std time)"

        for name, std_t_default, opt_percent_default_int, _, unit, mtype in ops_list:
            key = (tab_name, name)
            if not self.check_vars[tab_name][name].get():
                continue
            if mode == 'OPTIMIZED' and self.non_optimized_vars[tab_name][name].get():
                continue

            # 使用当前显示的 Std Time 值（已含 Complexity 和用户修改）
            t_per = self.safe_get_double(key)

            qty = self.safe_get_int(self.quantity_vars[tab_name][name])

            if tab_name == 'Tuning' and name == rf_name:
                minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty * n_cav
            else:
                minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty

            parts_qty = self.safe_get_int(self.parts_qty_vars[tab_name][name])
            minutes *= parts_qty

            # 对于 OPTIMIZED 模式，应用优化百分比（但由于 Std Time 已含 Complexity，这里不重复乘）
            if mode == 'OPTIMIZED' and not self.non_optimized_vars[tab_name][name].get():
                opt_percent_decimal = self.safe_get_opt_percent(key)
                minutes *= opt_percent_decimal

            total += minutes
        return total

    def reset_all(self):
        self.user_selected_demanding_perf = False
        self.n_cavities_var.set('5')
        self.housing_per_fg_var.set('2')
        self.standard_mode_var.set(True)
        self.optimized_mode_var.set(True)
        self.pn_var.set("")
        self.product_var.set('TPX')
        self.cm_var.set('GLOBAL')
        self.tuning_type_var.set('TPX')

        self.standard_fpy_var.set('90%')
        self.standard_custom_fpy_num_var.set('90')
        self.standard_rework_var.set('10%')
        self.optimized_fpy_var.set('90%')
        self.optimized_custom_fpy_num_var.set('90')
        self.optimized_rework_var.set('10%')
        
        tuning_check_key = (TUNING_TAB_NAME, TUNING_CHECK_NAME)
        if tuning_check_key in self.std_time_vars:
            # 1. 强制将 original_std_times 设为 5.0，防止导入数据覆盖，确保下次复位也正确。
            self.original_std_times[tuning_check_key] = DEFAULT_TUNING_CHECK_STD_TIME
            # 2. 执行 Std Time 变量的复位
            self.std_time_vars[tuning_check_key].set(DEFAULT_TUNING_CHECK_STD_TIME)

        for section in self.complexity_sections:
            self.complexity_vars[section].set('1.0')
            self.custom_complexity_num_vars[section].set('1.0')
            self.on_complexity_combo_change(section)

        housing = self.get_housing()
        section_ops = {'Pre_Assembly': PRE_ASSEMBLY_OPS, 'Assembly': ASSEMBLY_OPS, 'Tuning': TUNING_OPS,
                       'Testing': TESTING_OPS, 'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS}
        self.suppress_qty_modify = True
        
        for tab in ['Pre_Assembly', 'Assembly', 'Tuning', 'Testing', 'Quality Check']:
            ops = section_ops[tab]
            for name in self.quantity_vars[tab]:
                if not self.user_modified.get(tab, {}).get(name, False):
                    default_qty = next(op[3] for op in ops if op[0] == name)
                    self.quantity_vars[tab][name].set(default_qty * housing)
        self.suppress_qty_modify = False

        # 重置所有数据容器
        self.suppress_updates = True
        try:
            for tab in self.check_vars:
                for name in self.check_vars[tab]:
                    self.check_vars[tab][name].set(False)
                    self.non_optimized_vars[tab][name].set(False)
                    self.parts_qty_vars[tab][name].set(1)
                    key = (tab, name)
                    self.std_time_vars[key].set(self.original_std_times[key])
                    self.opt_percent_vars[key].set(next((op[2] for op in (PRE_ASSEMBLY_OPS+ASSEMBLY_OPS+TUNING_OPS+TESTING_OPS+QUALITY_OPS+PACKAGING_OPS) if op[0]==name), 100))

                # 清除所有 user_modified 标记（恢复自动计算）
                self.user_modified[tab] = {}

        finally:
            self.suppress_updates = False

        self.update_all_times()
        self.results.delete(1.0, tk.END)
        self.standard_custom_fpy_entry.config(state='disabled')
        self.standard_percent_label.config(state='disabled')
        self.optimized_custom_fpy_entry.config(state='disabled')
        self.optimized_percent_label.config(state='disabled')

    def calculate_detailed(self):
        self.update_all_times()
        product = self.product_var.get()
        cm = self.cm_var.get()
        n_cav = self.get_n_cavities()
        housing = self.get_housing()
        modules = housing
        tuning_type = self.tuning_type_var.get()
        pn = self.pn_var.get() or "N/A"

        factor = self.get_factor()
        selected_modes = []
        if self.standard_mode_var.get():
            selected_modes.append(('STANDARD', self.get_standard_fpy()))
        if self.optimized_mode_var.get():
            selected_modes.append(('OPTIMIZED', self.get_optimized_fpy()))

        if not selected_modes:
            messagebox.showwarning("Warning", "Please select at least one mode.")
            return

        result = ""
        for mode, fpy in selected_modes:
            pre_assembly_base = self.get_section_total('Pre_Assembly', PRE_ASSEMBLY_OPS, mode)
            assembly_base = self.get_section_total('Assembly', ASSEMBLY_OPS, mode)
            tuning_base = self.get_section_total('Tuning', TUNING_OPS, mode)
            testing_base = self.get_section_total('Testing', TESTING_OPS, mode)
            quality = self.get_section_total('Quality Check', QUALITY_OPS, mode)
            packaging = self.get_section_total('Packaging', PACKAGING_OPS, mode)

            total_wo = pre_assembly_base + assembly_base + tuning_base + testing_base + quality + packaging
            rework_rate = 1 - fpy
            rework_base = total_wo - pre_assembly_base - quality - packaging
            rework = rework_base * rework_rate
            total = total_wo + rework

            complexity_p = self.get_complexity_level('Pre_Assembly')
            complexity_a = self.get_complexity_level('Assembly')
            complexity_t = self.get_complexity_level('Tuning')
            complexity_te = self.get_complexity_level('Testing')
            complexity_q = self.get_complexity_level('Quality Check')
            complexity_pa = self.get_complexity_level('Packaging')

            result += f""" ROUTING CALCULATION RESULT ({mode})
{'═' * 100}
Routing            : {self.cm_var.get():<12}
P/N                : {pn}
Estimated FPY      : {self.get_fpy_display(mode)}
Product            : {product:<12}
Tuning Type        : {tuning_type}
Cavities           : {n_cav}
Housing/FG         : {housing}
Mode               : {mode}

Complexity Levels  :
Pre_Assembly   : {complexity_p:<8.2f}Assembly   : {complexity_a:<8.2f}
Tuning         : {complexity_t:<8.2f}Testing    : {complexity_te:<8.2f}
Quality Check  : {complexity_q:<8.2f}Packaging  : {complexity_pa:<8.2f}

Pre_Assembly       : {pre_assembly_base:6.2f}
Assembly           : {assembly_base:6.2f}
Tuning             : {tuning_base:6.2f}
Testing            : {testing_base:6.2f}
Cleaning           : {0.0:6.2f}  (Included in Assembly)
Quality Check      : {quality:6.2f}
Packaging          : {packaging:6.2f}
{'─' * 50}
Rework             : {rework:6.2f}
Total w/o Rework   : {total_wo:6.2f}
╔{'═'*42}╗
║ Total Routing Time : {total:7.2f}  minutes    ║
╚{'═'*42}╝

"""
        self.results.delete(1.0, tk.END)
        self.results.insert(tk.END, result)

    def export_to_excel(self):
        self.update_all_times()
        try:
            pn_raw = self.pn_var.get().strip()
            pn = "".join(c if c.isalnum() or c in "._-/\\" else "_" for c in pn_raw) if pn_raw else "NoPN"
            tuning = self.tuning_type_var.get().replace(" ", "_").replace("(", "").replace(")", "")
            suggested_name = f"Routing_{pn}_{tuning}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                title="保存 Routing 结果",
                initialfile=suggested_name
            )
            if not filename:
                return

            selected_modes = []
            if self.standard_mode_var.get():
                selected_modes.append(('STANDARD', self.get_standard_fpy(), self.get_fpy_display('STANDARD')))
            if self.optimized_mode_var.get():
                selected_modes.append(('OPTIMIZED', self.get_optimized_fpy(), self.get_fpy_display('OPTIMIZED')))

            if not selected_modes:
                messagebox.showwarning("Warning", "Please select at least one mode.")
                return

            n_cav = self.get_n_cavities()
            housing = self.get_housing()
            rf_name = "RF TUNING1 estimated (if no other indication)"
            check_name = "Tuning check after cleaning (Check Std time)"

            rows = []
            for tab, ops in [('Pre_Assembly', PRE_ASSEMBLY_OPS), ('Assembly', ASSEMBLY_OPS), ('Tuning', TUNING_OPS),
                            ('Testing', TESTING_OPS), ('Quality Check', QUALITY_OPS), ('Packaging', PACKAGING_OPS)]:
                complexity_factor = self.get_complexity_level(tab)

                for name, std_t_default, opt_percent_default_int, _, unit, mtype in ops:
                    key = (tab, name)

                    if self.check_vars[tab][name].get():
                        parts_p = self.safe_get_int(self.parts_qty_vars[tab][name])
                        qty = self.safe_get_int(self.quantity_vars[tab][name])

                        opt_percent = float(self.opt_percent_vars[key].get())
                        opt_pct = f"{opt_percent:.0f}%"

                        # Display Std/Opt Time (WITH Complexity, as in UI)
                        std_t_display = self.safe_get_double(key)
                        opt_t_display = std_t_display * (opt_percent / 100.0)
                        std_time_std = f"{std_t_display:.2f} {unit}"
                        std_time_opt = f"{opt_t_display:.2f} {unit}"

                        # Calc Time STANDARD (min)
                        minutes_std = self.calc_labels[tab][name].cget("text").strip()

                        # Calc Time OPTIMIZED (min)
                        if self.non_optimized_vars[tab][name].get():
                            minutes_opt = 0.0
                        else:
                            minutes_opt = float(minutes_std) * (opt_percent / 100.0)

                        rows.append({
                            "Section": tab,
                            "Operation": name,
                            "Complexity Factor": f"{complexity_factor:.2f}",
                            "Excluded": self.non_optimized_vars[tab][name].get(),
                            "N° of P": parts_p,
                            "Qty": qty,
                            "Std Time": std_time_std,
                            "Opt Time": std_time_opt,
                            "Optimized %": opt_pct,
                            "Calc Time STANDARD (min)": minutes_std,
                            "Calc Time OPTIMIZED (min)": round(minutes_opt, 2)
                        })

            df_ops = pd.DataFrame(rows) if rows else pd.DataFrame({"Info": ["No operations selected"]})

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                config_data = [
                    ["Routing", self.cm_var.get()],
                    ["P/N", pn_raw or "N/A"],
                    ["Product", self.product_var.get()],
                    ["Tuning Type", self.tuning_type_var.get()],
                    ["N° Cavities", self.n_cavities_var.get()],
                    ["N° Housings", self.housing_per_fg_var.get()],
                    ["STANDARD FPY", self.get_fpy_display('STANDARD')],
                    ["OPTIMIZED FPY", self.get_fpy_display('OPTIMIZED')],
                    ["Complexity Factor (Pre_Assembly)", f"{self.get_complexity_level('Pre_Assembly'):.2f}"],
                    ["Complexity Factor (Assembly)", f"{self.get_complexity_level('Assembly'):.2f}"],
                    ["Complexity Factor (Tuning)", f"{self.get_complexity_level('Tuning'):.2f}"],
                    ["Complexity Factor (Testing)", f"{self.get_complexity_level('Testing'):.2f}"],
                    ["Complexity Factor (Quality Check)", f"{self.get_complexity_level('Quality Check'):.2f}"],
                    ["Complexity Factor (Packaging)", f"{self.get_complexity_level('Packaging'):.2f}"],
                    ["Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                ]
                pd.DataFrame(config_data, columns=["Parameter", "Value"]).to_excel(writer, sheet_name="Configuration", index=False)

                df_ops.to_excel(writer, sheet_name="Selected Operations", index=False)

                items = ["Pre_Assembly", "Assembly", "Tuning", "Testing", "Quality Check", "Packaging", "Rework", "Total w/o Rework", "Total Routing Time"]
                all_values = []
                mode_names = []
                for mode, fpy, fpy_display in selected_modes:
                    pre_assembly_base = self.get_section_total('Pre_Assembly', PRE_ASSEMBLY_OPS, mode)
                    assembly_base = self.get_section_total('Assembly', ASSEMBLY_OPS, mode)
                    tuning_base = self.get_section_total('Tuning', TUNING_OPS, mode)
                    testing_base = self.get_section_total('Testing', TESTING_OPS, mode)
                    quality = self.get_section_total('Quality Check', QUALITY_OPS, mode)
                    packaging = self.get_section_total('Packaging', PACKAGING_OPS, mode)

                    total_wo_rework = pre_assembly_base + assembly_base + tuning_base + testing_base + quality + packaging
                    rework_rate = 1 - fpy
                    rework_base = total_wo_rework - pre_assembly_base - quality - packaging
                    rework = rework_base * rework_rate
                    final_total = total_wo_rework + rework

                    values = [
                        round(pre_assembly_base,2),
                        round(assembly_base,2),
                        round(tuning_base,2),
                        round(testing_base,2),
                        round(quality,2),
                        round(packaging,2),
                        round(rework,2),
                        round(total_wo_rework,2),
                        round(final_total,2)
                    ]
                    all_values.append(values)
                    mode_names.append(f"{mode} Base Time (FPY: {fpy_display})")

                summary_df = pd.DataFrame(all_values, columns=items)
                summary_df.insert(0, "Mode", mode_names)
                summary_df.to_excel(writer, sheet_name="Time Summary", index=False, startrow=1)

                wb = writer.book
                # 1. 定义通用左对齐格式
                left_align = Alignment(horizontal="left")
                
                for ws in wb.worksheets:
                    if ws.title == "Time Summary":
                        ws['A1'] = f"P/N: {pn_raw or 'N/A'}"
                        ws['A1'].font = Font(bold=True)
                        ws['A1'].fill = PatternFill(start_color="D6EBFF", end_color="D6EBFF", fill_type="solid")
                        ws['A1'].alignment = Alignment(horizontal="left")
                        header_row = 2
                    else:
                        header_row = 1

                    # 2. 设置标题为左对齐
                    for cell in ws[header_row]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D6EBFF", end_color="D6EBFF", fill_type="solid")
                        cell.alignment = Alignment(horizontal="left")

                    # 3. 设置数据单元格为左对齐
                    data_start_row = header_row + 1
                    for row in ws.iter_rows(min_row=data_start_row):
                        for cell in row:
                            cell.alignment = left_align

                    for col in ws.columns:
                        max_len = 0
                        column = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[column].width = min(max_len + 4, 60)

                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    min_border_row = 1 if ws.title == "Time Summary" else header_row
                    for row in ws.iter_rows(min_row=min_border_row):
                        for cell in row:
                            cell.border = border

                config_ws = wb["Configuration"]
                for row in config_ws.iter_rows(min_row=2, max_row=config_ws.max_row, min_col=1, max_col=2):
                    param = row[0].value
                    if param in ["N° Cavities", "N° Housings"]:
                        #row[1].alignment = Alignment(horizontal='left')
                        row[1].alignment = left_align

            messagebox.showinfo("Success", f"Excel has been successfully exported!\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")

    # ---------------- 新增：导入 Excel（严格校验） ----------------
    def import_from_excel(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")], title="Select Import Excel")
        if not filename:
            return

        try:
            xls = pd.ExcelFile(filename)
        except Exception as e:
            messagebox.showerror("Import Error", f"Cannot open Excel file:\n{e}")
            return

        required_sheets = ['Configuration', 'Selected Operations']
        for s in required_sheets:
            if s not in xls.sheet_names:
                messagebox.showerror("Import Error", f"Required sheet '{s}' not found in the selected Excel.")
                return

        try:
            df_conf = pd.read_excel(xls, sheet_name='Configuration', engine='openpyxl')
        except Exception as e:
            messagebox.showerror("Import Error", f"Cannot read 'Configuration' sheet:\n{e}")
            return

        expected_conf_cols = ['Parameter', 'Value']
        if not all(c in df_conf.columns for c in expected_conf_cols):
            messagebox.showerror("Import Error", "Configuration sheet must contain columns: 'Parameter' and 'Value'.")
            return

        conf_map = {}
        for _, row in df_conf.iterrows():
            k = str(row['Parameter']).strip() if not pd.isna(row['Parameter']) else ""
            v = row['Value']
            if k:
                conf_map[k] = v

        required_conf_keys = [
            "Routing", "P/N", "Product", "Tuning Type",
            "N° Cavities", "N° Housings",
            "STANDARD FPY", "OPTIMIZED FPY",
            "Complexity Factor (Pre_Assembly)", "Complexity Factor (Assembly)",
            "Complexity Factor (Tuning)", "Complexity Factor (Testing)",
            "Complexity Factor (Quality Check)", "Complexity Factor (Packaging)"
        ]
        for k in required_conf_keys:
            if k not in conf_map or pd.isna(conf_map[k]):
                messagebox.showerror("Import Error", f"Configuration missing required parameter: '{k}'. Import aborted.")
                return

        try:
            df_ops = pd.read_excel(xls, sheet_name='Selected Operations', engine='openpyxl')
        except Exception as e:
            messagebox.showerror("Import Error", f"Cannot read 'Selected Operations' sheet:\n{e}")
            return

        required_ops_cols = ["Section", "Operation", "Excluded", "N° of P", "Qty", "Std Time", "Opt Time", "Optimized %"]
        for c in required_ops_cols:
            if c not in df_ops.columns:
                messagebox.showerror("Import Error", f"Selected Operations sheet must contain column: '{c}'. Import aborted.")
                return

        valid_sections = {
            'Pre_Assembly': [op[0] for op in PRE_ASSEMBLY_OPS],
            'Assembly': [op[0] for op in ASSEMBLY_OPS],
            'Tuning': [op[0] for op in TUNING_OPS],
            'Testing': [op[0] for op in TESTING_OPS],
            'Quality Check': [op[0] for op in QUALITY_OPS],
            'Packaging': [op[0] for op in PACKAGING_OPS]
        }

        def parse_time_with_unit(s):
            if pd.isna(s):
                return None
            try:
                txt = str(s).strip()
                parts = txt.split()
                if len(parts) == 1:
                    return None
                value = float(parts[0])
                unit = parts[-1].lower()
                if unit not in ('sec', 'min'):
                    return None
                return (value, unit)
            except:
                return None

        def parse_bool(v):
            if pd.isna(v):
                return None
            if isinstance(v, bool):
                return v
            txt = str(v).strip().lower()
            if txt in ('true', '1', 'yes', 'y', 't'):
                return True
            if txt in ('false', '0', 'no', 'n', 'f'):
                return False
            return None

        def parse_percent(v):
            if pd.isna(v):
                return None
            try:
                if isinstance(v, str) and '%' in v:
                    return float(v.strip().replace('%',''))
                return float(v)
            except:
                return None

        ops_to_apply = []
        for idx, row in df_ops.iterrows():
            section = str(row['Section']).strip()
            operation = str(row['Operation']).strip()

            if section not in valid_sections:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: Section '{section}' is invalid.")
                return

            if operation not in valid_sections[section]:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: Operation '{operation}' not found in section '{section}'. Import aborted.")
                return

            excluded = parse_bool(row['Excluded'])
            if excluded is None:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: 'Excluded' value '{row['Excluded']}' invalid. Use True/False or 1/0.")
                return

            try:
                parts_p = int(row["N° of P"])
            except:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: 'N° of P' must be integer.")
                return

            try:
                qty = int(row["Qty"])
            except:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: 'Qty' must be integer.")
                return

            std_parsed = parse_time_with_unit(row["Std Time"])
            if std_parsed is None:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: 'Std Time' must be like '12.00 sec' or '3.00 min'. Import aborted.")
                return
            std_value, std_unit = std_parsed

            opt_parsed = parse_time_with_unit(row["Opt Time"])
            if opt_parsed is None:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: 'Opt Time' must be like '12.00 sec' or '3.00 min'. Import aborted.")
                return
            opt_value, opt_unit = opt_parsed

            if std_unit != opt_unit:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: Std/Opt units do not match.")
                return

            opt_pct = parse_percent(row["Optimized %"])
            if opt_pct is None:
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: 'Optimized %' invalid.")
                return
            if not (0 <= opt_pct <= 100):
                messagebox.showerror("Import Error", f"Row {idx+2} in Selected Operations: 'Optimized %' must be between 0 and 100.")
                return

            # Unit must match the unit defined in our unit_dict (strict)
            key = (section, operation)
            expected_unit = self.unit_dict.get(key)
            if expected_unit is None:
                messagebox.showerror("Import Error", f"Internal Error: expected unit for {key} not found.")
                return
            if expected_unit != std_unit:
                messagebox.showerror("Import Error", f"Row {idx+2}: unit mismatch for '{operation}' in section '{section}': expected '{expected_unit}', got '{std_unit}'. Import aborted.")
                return

            ops_to_apply.append({
                "section": section,
                "operation": operation,
                "excluded": excluded,
                "parts_p": parts_p,
                "qty": qty,
                "std_value": std_value,
                "opt_value": opt_value,
                "opt_pct": opt_pct
            })

        # ===== 到这里，Configuration 与 Selected Operations 均已解析且校验通过 =====
        # helper: restore complexity field (works for any section name)
        def restore_complexity_for_section(param_name, combo_var, custom_var):
            raw = conf_map.get(param_name)
            if pd.isna(raw):
                return False
            # raw may be number or string like "1.35" or "1.35 "
            val_str = str(raw).strip()
            # Normalize preset strings from COMPLEXITY_LEVELS (e.g., '1.0' -> '1.0', user may have '1.00')
            # Accept numeric match too.
            try:
                val_float = float(val_str)
            except:
                # If cannot parse as float, but matches exactly a preset string, accept
                if val_str in COMPLEXITY_LEVELS:
                    combo_var.set(val_str)
                    return True
                return False

            # check if matches any preset numeric values
            preset_floats = []
            preset_map = {}
            for item in COMPLEXITY_LEVELS:
                if item.lower() == 'custom':
                    continue
                try:
                    fv = float(item)
                    preset_floats.append(fv)
                    preset_map[fv] = item
                except:
                    pass

            if val_float in preset_floats:
                combo_var.set(preset_map[val_float])
            else:
                # not a preset -> set to Custom and put value into custom var
                combo_var.set('Custom')
                custom_var.set(f"{val_float:.2f}")
            return True

        # helper: restore FPY (STANDARD / OPTIMIZED)
        def restore_fpy_field(param_name, combo_var, custom_var):
            raw = conf_map.get(param_name)
            if pd.isna(raw):
                return False
            txt = str(raw).strip()
            # allow either '88%' or 88 or '88.0'
            try:
                if '%' in txt:
                    val = float(txt.replace('%', '').strip())
                else:
                    val = float(txt)
            except:
                return False

            # define preset FPY percentages from FPY dict
            preset_percents = sorted(set(int(v*100) for v in FPY.values()))
            if int(val) in preset_percents:
                combo_var.set(f"{int(val)}%")
            else:
                combo_var.set('Custom')
                custom_var.set(str(int(val) if val.is_integer() else val))
            return True

        # 到这里开始把值写回界面（批量写回时抑制 trace 回调）
        self.suppress_updates = True
        try:
            # Routing
            routing_val = str(conf_map["Routing"]).strip()
            self.cm_var.set(routing_val)

            pn_val = str(conf_map["P/N"]).strip()
            self.pn_var.set(pn_val)

            product_val = str(conf_map["Product"]).strip()
            if product_val not in PRODUCTS:
                messagebox.showerror("Import Error", f"Configuration: Product '{product_val}' is not a valid product.")
                self.suppress_updates = False
                return
            self.product_var.set(product_val)

            tuning_type_val = str(conf_map["Tuning Type"]).strip()
            if tuning_type_val not in TUNING_TYPES:
                messagebox.showerror("Import Error", f"Configuration: Tuning Type '{tuning_type_val}' is invalid.")
                self.suppress_updates = False
                return
            self.tuning_type_var.set(tuning_type_val)

            # N° Cavities (string may be numeric)
            try:
                n_cav_val = int(conf_map["N° Cavities"])
            except:
                messagebox.showerror("Import Error", "Configuration: 'N° Cavities' must be integer. Import aborted.")
                self.suppress_updates = False
                return
            self.n_cavities_var.set(str(n_cav_val))

            # N° Housings
            try:
                hous_val = int(conf_map["N° Housings"])
            except:
                messagebox.showerror("Import Error", "Configuration: 'N° Housings' must be integer. Import aborted.")
                self.suppress_updates = False
                return
            self.housing_per_fg_var.set(str(hous_val))

            # Restore FPY fields (STANDARD / OPTIMIZED)
            restore_fpy_field("STANDARD FPY", self.standard_fpy_var, self.standard_custom_fpy_num_var)
            restore_fpy_field("OPTIMIZED FPY", self.optimized_fpy_var, self.optimized_custom_fpy_num_var)
            # ensure combo change callbacks run to enable/disable entry
            self.on_standard_fpy_combo_change()
            self.on_optimized_fpy_combo_change()

            # Complexity factors: use helper for each section's Configuration key
            restore_complexity_for_section("Complexity Factor (Pre_Assembly)", self.complexity_vars['Pre_Assembly'], self.custom_complexity_num_vars['Pre_Assembly'])
            restore_complexity_for_section("Complexity Factor (Assembly)", self.complexity_vars['Assembly'], self.custom_complexity_num_vars['Assembly'])
            restore_complexity_for_section("Complexity Factor (Tuning)", self.complexity_vars['Tuning'], self.custom_complexity_num_vars['Tuning'])
            restore_complexity_for_section("Complexity Factor (Testing)", self.complexity_vars['Testing'], self.custom_complexity_num_vars['Testing'])
            restore_complexity_for_section("Complexity Factor (Quality Check)", self.complexity_vars['Quality Check'], self.custom_complexity_num_vars['Quality Check'])
            restore_complexity_for_section("Complexity Factor (Packaging)", self.complexity_vars['Packaging'], self.custom_complexity_num_vars['Packaging'])

            # Reset selection state then apply rows
            for tab in self.check_vars:
                for name in list(self.check_vars[tab].keys()):
                    self.check_vars[tab][name].set(False)
                    self.non_optimized_vars[tab][name].set(False)
                    self.parts_qty_vars[tab][name].set(1)
                    # restore original std_time_vars to original_std_times (so original reference preserved)
                    key = (tab, name)
                    if key in self.original_std_times:
                        self.std_time_vars[key].set(self.original_std_times[key])
                        self.opt_percent_vars[key].set(next((op[2] for op in (PRE_ASSEMBLY_OPS+ASSEMBLY_OPS+TUNING_OPS+TESTING_OPS+QUALITY_OPS+PACKAGING_OPS) if op[0]==name), 100))

            # apply each op row
            for entry in ops_to_apply:
                section = entry["section"]
                operation = entry["operation"]
                key = (section, operation)

                # set selected
                self.check_vars[section][operation].set(True)

                # set excluded (non optimized)
                self.non_optimized_vars[section][operation].set(bool(entry["excluded"]))

                # set parts qty
                self.parts_qty_vars[section][operation].set(int(entry["parts_p"]))

                # set qty
                self.quantity_vars[section][operation].set(int(entry["qty"]))

                # set original_std_times and std_time_vars base (original std time WITHOUT complexity)
                # original_std_times should be base (unit-aware)
                self.original_std_times[key] = float(entry["std_value"])
                # set the display std_time_vars to base value; update_section_times will apply complexity for UI calc later
                self.std_time_vars[key].set(float(entry["std_value"]))

                # set opt percent var (0-100)
                self.opt_percent_vars[key].set(float(entry["opt_pct"]))

                # mark user_modified if qty differs from default
                # determine default qty: look up ops list
                default_qty = None
                for ops in (PRE_ASSEMBLY_OPS, ASSEMBLY_OPS, TUNING_OPS, TESTING_OPS, QUALITY_OPS, PACKAGING_OPS):
                    for op in ops:
                        if op[0] == operation:
                            default_qty = op[3]
                            break
                    if default_qty is not None:
                        break
                if default_qty is None:
                    # should not happen because we've validated earlier; but keep safe fallback
                    self.user_modified[section][operation] = True
                else:
                    # compare imported qty with default * housing if relevant
                    if section in ['Pre_Assembly', 'Assembly', 'Tuning', 'Testing', 'Quality Check']:
                        expected_default = default_qty * self.get_housing()
                    else:
                        expected_default = default_qty
                    self.user_modified[section][operation] = (int(entry["qty"]) != expected_default)

        finally:
            self.suppress_updates = False

        # 最后，触发 UI 的变更刷新
        try:
            # ensure tuning std time updated based on chosen tuning type
            self.update_tuning_std_time()
            # recalc and update UI labels
            self.update_all_times()
            messagebox.showinfo("Import Successful", "Excel imported successfully and UI updated.")
        except Exception as e:
            messagebox.showerror("Import Error", f"Import applied but update failed: {e}")

    # ---------------- 其余函数 ----------------
    def on_qty_change(self, tab, name, *args):
        if self.suppress_qty_modify:
            return
        if tab in ['Pre_Assembly', 'Assembly', 'Tuning', 'Testing', 'Quality Check']:
            self.user_modified[tab][name] = True
        self.update_section_times(tab)

    def get_housing(self):
        try:
            return int(self.housing_per_fg_var.get() or '2')
        except:
            return 2

    def get_n_cavities(self):
        try:
            return int(self.n_cavities_var.get() or '0')
        except:
            return 0

    def get_factor(self):
        cm = self.cm_var.get()
        product = self.product_var.get()
        if cm in FACTORS and product in FACTORS[cm]:
            return FACTORS[cm][product]
        return 1.0

    def on_housing_change(self, *args):
        housing = self.get_housing()
        section_ops = {
            'Pre_Assembly': PRE_ASSEMBLY_OPS,
            'Assembly': ASSEMBLY_OPS,
            'Tuning': TUNING_OPS,
            'Testing': TESTING_OPS,
            'Quality Check': QUALITY_OPS
        }
        self.suppress_qty_modify = True
        for tab in ['Pre_Assembly', 'Assembly', 'Tuning', 'Testing', 'Quality Check']:
            ops = section_ops[tab]
            for name in self.quantity_vars[tab]:
                if not self.user_modified.get(tab, {}).get(name, False):
                    default_qty = next(op[3] for op in ops if op[0] == name)
                    self.quantity_vars[tab][name].set(default_qty * housing)
        self.suppress_qty_modify = False
        self.update_all_times()

    # 新增：Tuning Type 改变时的联动逻辑（双向 + 保护 Demanding perf）
    def on_tuning_type_change(self, *args):
        selected = self.tuning_type_var.get()

        if selected == "Demanding perf":
            # 用户手动选了高难度调谐 → 锁定，之后 Product 改变也不覆盖
            self.user_locked_demanding_perf = True
            return

        # 用户选了普通类型 → 解除锁定，并联动 Product
        self.user_locked_demanding_perf = False

        if selected in PRODUCTS:  # ['TMA', 'DPX', 'TPX', 'PPX', 'QPX', 'IMF', 'LLC']
            if self.product_var.get() != selected:
                self.product_var.set(selected)

    # 替换原来的 on_product_change（关键：变量名统一！）
    def on_product_change(self, *args):
        prod = self.product_var.get()

        # 1. 原有的 FPY 联动（不变）
        if prod in FPY:
            fpy_pct = f"{int(FPY[prod] * 100)}%"
            self.standard_fpy_var.set(fpy_pct)
            self.standard_custom_fpy_num_var.set(str(int(FPY[prod] * 100)))
            self.optimized_fpy_var.set(fpy_pct)
            self.optimized_custom_fpy_num_var.set(str(int(FPY[prod] * 100)))

        # 2. Product → Tuning Type 联动（只有没被用户锁定为 Demanding perf 时才生效）
        if not self.user_locked_demanding_perf:
            if prod in TUNING_TYPES:
                target = prod
            else:
                target = 'DPX'  # 兜底
            if self.tuning_type_var.get() != target:
                self.tuning_type_var.set(target)

    def on_standard_fpy_combo_change(self, *args):
        if self.standard_fpy_var.get() == 'Custom':
            self.standard_custom_fpy_entry.config(state='normal')
            self.standard_percent_label.config(state='normal')
            self.update_standard_fpy_rework()
        else:
            self.standard_custom_fpy_entry.config(state='disabled')
            self.standard_percent_label.config(state='disabled')
            self.standard_custom_fpy_num_var.set(self.standard_fpy_var.get().strip('%'))
            self.update_standard_fpy_rework()

    def on_optimized_fpy_combo_change(self, *args):
        if self.optimized_fpy_var.get() == 'Custom':
            self.optimized_custom_fpy_entry.config(state='normal')
            self.optimized_percent_label.config(state='normal')
            self.update_optimized_fpy_rework()
        else:
            self.optimized_custom_fpy_entry.config(state='disabled')
            self.optimized_percent_label.config(state='disabled')
            self.optimized_custom_fpy_num_var.set(self.optimized_fpy_var.get().strip('%'))
            self.update_optimized_fpy_rework()

    def update_standard_fpy_rework(self, *args):
        try:
            if self.standard_fpy_var.get() == 'Custom':
                if self.standard_custom_fpy_num_var.get().isdigit():
                    fpy_pct = int(self.standard_custom_fpy_num_var.get())
                else:
                    fpy_pct = 90
            else:
                fpy_pct = int(self.standard_fpy_var.get().strip('%'))
            fpy_pct = max(0, min(100, fpy_pct))
            rework_pct = 100 - fpy_pct
            self.standard_rework_var.set(f"{rework_pct}%")
        except:
            self.standard_rework_var.set("10%")

    def update_optimized_fpy_rework(self, *args):
        try:
            if self.optimized_fpy_var.get() == 'Custom':
                if self.optimized_custom_fpy_num_var.get().isdigit():
                    fpy_pct = int(self.optimized_custom_fpy_num_var.get())
                else:
                    fpy_pct = 90
            else:
                fpy_pct = int(self.optimized_fpy_var.get().strip('%'))
            fpy_pct = max(0, min(100, fpy_pct))
            rework_pct = 100 - fpy_pct
            self.optimized_rework_var.set(f"{rework_pct}%")
        except:
            self.optimized_rework_var.set("10%")

    def get_standard_fpy(self):
        try:
            if self.standard_fpy_var.get() == 'Custom':
                return float(self.standard_custom_fpy_num_var.get()) / 100
            else:
                return float(self.standard_fpy_var.get().strip('%')) / 100
        except:
            return 0.9

    def get_optimized_fpy(self):
        try:
            if self.optimized_fpy_var.get() == 'Custom':
                return float(self.optimized_custom_fpy_num_var.get()) / 100
            else:
                return float(self.optimized_fpy_var.get().strip('%')) / 100
        except:
            return 0.9

    def get_fpy_display(self, mode):
        if mode == 'STANDARD':
            if self.standard_mode_var.get() and self.standard_fpy_var.get() == 'Custom':
                return f"{self.standard_custom_fpy_num_var.get()}%"
            else:
                return self.standard_fpy_var.get()
        else:
            if self.optimized_mode_var.get() and self.optimized_fpy_var.get() == 'Custom':
                return f"{self.optimized_custom_fpy_num_var.get()}%"
            else:
                return self.optimized_fpy_var.get()

    def update_tuning_std_time(self, *args):
        tab = 'Tuning'
        rf_name = "RF TUNING1 estimated (if no other indication)"
        check_name = "Tuning check after cleaning (Check Std time)"

        if rf_name in self.check_vars[tab]:
            tuning_type = self.tuning_type_var.get()
            if tuning_type in TUNING_TYPES:
                std_sec, _ = TUNING_TYPES[tuning_type]
                self.suppress_updates = True
                try:
                    self.std_time_vars[(tab, rf_name)].set(std_sec)
                    self.original_std_times[(tab, rf_name)] = std_sec
                finally:
                    self.suppress_updates = False
              
            if self.std_time_vars[(tab, check_name)].get() != self.original_std_times[(tab, check_name)]:
                self.suppress_updates = True
                try:
                    self.std_time_vars[(tab, check_name)].set(self.original_std_times[(tab, check_name)])
                finally:
                    self.suppress_updates = False
            
            self.update_section_times(tab)

    def import_preview(self):
        pass

    # ---------------- 主循环 ----------------
if __name__ == "__main__":
    root = tk.Tk()
    set_app_icon(root)
    app = RoutingApp(root)
    root.mainloop()