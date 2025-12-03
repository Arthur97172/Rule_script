# -*- coding: utf-8 -*-
"""
Routing Estimation Tool - 工业级终极完整版
功能：
1. 新增 P/N 输入框
2. 导出文件名自动包含 P/N + Tuning Type
3. 一键导出美观 Excel（3个Sheet）
4. 完整工序数据 + 实时计算 + Reset All
5. Qty 默认根据 Housing per FG 自动更新（用户修改后保持不变）
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import scrolledtext
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==================== 新增：用于 PyInstaller 打包和图标设置 ====================
import sys
import os

# ----------------------------------------------------
# [新增] PyInstaller 资源路径解析函数 (修复 onefile 模式路径问题)
# ----------------------------------------------------
def resource_path(relative_path):
    """
    获取资源文件的绝对路径。
    在开发模式下，返回相对路径。
    在 PyInstaller 打包模式下，返回临时解压路径 (_MEIPASS)。
    """
    try:
        # PyInstaller 打包后的临时目录
        base_path = sys._MEIPASS
    except Exception:
        # 开发模式下的当前脚本目录
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)
# ----------------------------------------------------

# ----------------------------------------------------
# 统一设置程序图标
# ----------------------------------------------------
def set_app_icon(window):
    """统一设置窗口图标，自动处理打包路径（静默失败，不输出任何信息）"""
    try:
        # 使用 resource_path 来获取 time.ico 的路径
        icon_path = resource_path("time.ico") 
        # ico 文件需要在打包时通过 --add-data 参数包含
        # PyInstaller 命令示例: pyinstaller --onefile --add-data "time.ico;."
        window.iconbitmap(icon_path)
    except Exception as e:
        # print(f"Icon loading error: {e}") # 调试时可打开
        pass  # 图标缺失时静默跳过，不影响程序运行

# ==============================================================================

# ================================= 数据定义 =================================

PRODUCTS = ['TMA', 'DPX', 'TPX', 'PPX', 'QPX', 'IMF', 'LLC']

FACTORS = {
    'TIGER': {
        'TMA': 1.0, 'DPX': 1.1, 'TPX': 1.28, 'PPX': 0.85, 'QPX': 0.95,
        'IMF': 0.83, 'LLC': 0.69
    },
    'TATFOOK': {
        'TMA': 1.39, 'DPX': 1.39, 'TPX': 1.43, 'PPX': 1.04, 'QPX': 0.92,
        'IMF': 1.02, 'LLC': 0.84
    },
    'LUXSHARE': {
        'TMA': 1.24, 'DPX': 1.35, 'TPX': 1.68, 'PPX': 1.18, 'QPX': 0.84,
        'IMF': 1.04, 'LLC': 0.87
    }
}

FPY = {'DPX': 0.95, 'TPX': 0.88, 'QPX': 0.85, 'TMA': 0.9, 'IMF': 0.85, 'LLC': 0.85, 'PPX': 0.85}

PRE_ASSEMBLY_OPS = [
    ("Tuning screw glued to tuners or plug", 12, 12, 1, 'sec', 'housing'),
    ("Lid cover+tuning screw+nut+washer", 8, 8, 1, 'sec', 'housing'),
    ("Connector+conductor soldered", 45, 45, 1, 'sec', 'housing'),
    ("Connector+conductor screwed", 5, 5, 1, 'sec', 'housing'),
    ("Other Pre_Assembly item", 10, 10, 1, 'sec', 'housing'),    
]

ASSEMBLY_OPS = [
    ("Low pass(or bias tee or coils) + thermalshrinking", 120, 120, 1, 'sec', 'housing'),
    ("Bias coil group", 30, 30, 1, 'sec', 'housing'),
    ("Lid cover + pin", 30, 30, 1, 'sec', 'housing'),
    ("Lid cover + line screwed", 5, 5, 1, 'sec', 'housing'),
    ("Lid cover + line soldered", 67, 67, 1, 'sec', 'housing'),
    ("Lid cover + connector", 15, 15, 1, 'sec', 'housing'),
    ("Lid cover + connector press in", 60, 60, 1, 'sec', 'housing'),
    ("Label for s/n", 5, 5, 1, 'sec', 'housing'),
    ("Housing + pin", 10, 10, 1, 'sec', 'housing'),
    ("Housing + connector flanged", 45, 45, 1, 'sec', 'housing'),
    ("Housing + connector screwed", 30, 30, 1, 'sec', 'housing'),
    ("Housing + junction screwed (time x 2 screws)", 10, 10, 1, 'sec', 'housing'),
    ("Housing + junction soldered by SOLDER STATION", 30, 30, 1, 'sec', 'housing'),
    ("Housing + junction soldered INDUCTION", 55, 55, 1, 'sec', 'housing'),
    ("Housing + line soldered by SOLDER STATION (complex)", 150, 150, 1, 'sec', 'housing'),
    ("Housing + line soldered INDUCTION (complex)", 150, 150, 1, 'sec', 'housing'),
    ("Housing + low pass(or bias tee)", 10, 10, 1, 'sec', 'housing'),
    ("Ret connector", 20, 20, 1, 'sec', 'housing'),
    ("Resonator", 7, 7, 1, 'sec', 'housing'),
    ("Housing + probe (including support fixing)", 9, 9, 1, 'sec', 'housing'),
    ("Apply solder paste for LID SOLDERING process", 300, 300, 1, 'sec', 'housing'),
    ("Tuning lid cover screwing", 3, 2.55, 1, 'sec', 'housing'),
    ("PWBA installation (including soldering)", 8, 8, 1, 'sec', 'housing'),
    ("Cable sub-assembly(or LED)", 15, 15, 1, 'sec', 'housing'),
    ("Outer cover (time x screw)", 4, 4, 1, 'sec', 'housing'),
    ("PWBA+Nickel Silver Cover placing", 3, 3, 1, 'sec', 'housing'),
    ("IP outer label ASSEMBLY", 60, 60, 1, 'sec', 'housing'),
    ("Housing + venting membrane", 10, 10, 1, 'sec', 'housing'),
    ("Housing + bracket +screw + nut + washer", 120, 120, 1, 'sec', 'housing'),
    ("Tuning lid open and close (time x screw)", 4, 4, 1, 'sec', 'housing'),
    ("Cleaning", 4.0, 4.0, 1, 'min', 'housing'),
    ("Other Assembly item", 5.0, 5.0, 1, 'min', 'housing'),    
]

TUNING_OPS = [
    ("RF TUNING1 estimated (if no other indication)", 108, 108, 1, 'sec', 'tuning'),  
    ("Tuning check after cleaning", 5.0, 5.0, 1, 'min', 'tuning'),
    ("TUNING 2 (E35)", 8.0, 8.0, 1, 'min', 'tuning'),
    ("Other Tuning item", 10.0, 10.0, 1, 'min', 'tuning'),    
]

TESTING_OPS = [
    ("PCBA programming (FW uploading)", 1.0, 1.0, 1, 'min', 'none'),
    ("NOISE FIGURE(mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("IP3(optional for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("AC/DC or Generic Measures", 3.0, 3.0, 1, 'min', 'none'),
    ("Fixed Tone IMD", 1.5, 1.5, 1, 'min', 'none'),
    ("Swept Tone IMD", 1.25, 1.25, 1, 'min', 'none'), 
    ("LINEARIZATION (mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("PID(mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("TDR on TMS (FINAL TEST)", 3.0, 2.7, 1, 'min', 'none'),
    ("PRESSURE TEST (Std)", 2.0, 2.0, 1, 'min', 'none'),
    ("PRESSURE TEST (IP Label)", 4.0, 4.0, 1, 'min', 'none'),    
    ("LED", 2.0, 2.0, 1, 'min', 'none'),
    ("Other Testing item", 5.0, 5.0, 1, 'min', 'none'),    
]

QUALITY_OPS = [
    ("Final check 100% (VISUAL INSPECTION)", 1.0, 1.0, 1, 'min', 'none'),
    ("Other Quality check item", 1.0, 1.0, 1, 'min', 'none'),
]
PACKAGING_OPS = [
    ("Packaging", 3.2, 3.2, 1, 'min', 'none'),
    ("Other Packaging item", 1.0, 1.0, 1, 'min', 'none'),  
]

TUNING_TYPES = {
    'DPX': (80, 72),
    'TPX': (90, 81),
    'PPX': (100, 90),
    'QPX': (100, 90),
    'TMA': (100, 90),
    'LLC': (120, 108),
    'IMF': (120, 108),
    'Demanding perf': (120, 108),
}

# ==============================================================================

class RoutingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Routing Estimation Tool - V1.0")
        self.root.geometry("1080x900")
        self.root.minsize(900, 720)

        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill='both', expand=True, padx=15, pady=12)

        # 新增 P/N
        self.pn_var = tk.StringVar(value="")

        self.product_var = tk.StringVar(value='TPX')
        self.cm_var = tk.StringVar(value='TIGER')
        self.standard_mode_var = tk.BooleanVar(value=True)
        self.optimized_mode_var = tk.BooleanVar(value=False)
        self.standard_fpy_var = tk.StringVar(value='88%')
        self.standard_custom_fpy_var = tk.StringVar(value='88%')
        self.standard_rework_var = tk.StringVar(value='12%')
        self.optimized_fpy_var = tk.StringVar(value='88%')
        self.optimized_custom_fpy_var = tk.StringVar(value='88%')
        self.optimized_rework_var = tk.StringVar(value='12%')
        self.n_cavities_var = tk.StringVar(value='0')
        self.housing_per_fg_var = tk.StringVar(value='2')
        self.tuning_type_var = tk.StringVar(value='DPX')

        self.suppress_qty_modify = False

        self.original_std_times = {}
        self.check_vars = {}
        self.quantity_vars = {}
        self.parts_qty_vars = {}
        self.std_time_vars = {}
        self.opt_time_dict = {}
        self.unit_dict = {}
        self.multi_type_dict = {}
        self.calc_labels = {}
        self.non_optimized_vars = {}

        # 新增：用户修改标记（用于 Qty 自动更新）
        self.user_modified = {}

        self.setup_data()
        self.setup_ui()

        self.standard_mode_var.trace_add('write', self.update_all_times)
        self.optimized_mode_var.trace_add('write', self.update_all_times)
        self.housing_per_fg_var.trace_add('write', self.on_housing_change)
        self.n_cavities_var.trace_add('write', self.update_all_times)
        self.product_var.trace_add('write', self.on_product_change)
        self.standard_fpy_var.trace_add('write', self.on_standard_fpy_combo_change)
        self.standard_custom_fpy_var.trace_add('write', self.update_standard_fpy_rework)
        self.optimized_fpy_var.trace_add('write', self.on_optimized_fpy_combo_change)
        self.optimized_custom_fpy_var.trace_add('write', self.update_optimized_fpy_rework)
        self.tuning_type_var.trace_add('write', self.update_tuning_std_time)
        self.cm_var.trace_add('write', self.update_all_times)

        self.update_all_times()
        self.on_product_change()
        self.update_standard_fpy_rework()
        self.update_optimized_fpy_rework()
        self.on_housing_change()
        self.update_tuning_std_time()

    def safe_get_int(self, var, default=1):
        try:
            return var.get()
        except tk.TclError:
            return default

    def safe_get_double(self, key):
        tab_name, name = key
        try:
            return self.std_time_vars[key].get()
        except tk.TclError:
            orig = self.original_std_times[key]
            self.std_time_vars[key].set(orig)
            return orig

    def setup_data(self):
        sections = {
            'Pre_Assembly': PRE_ASSEMBLY_OPS,
            'Assembly': ASSEMBLY_OPS,
            'Tuning': TUNING_OPS,
            'Testing': TESTING_OPS,
            'Quality Check': QUALITY_OPS,
            'Packaging': PACKAGING_OPS
        }
        for tab, ops in sections.items():
            self.check_vars[tab] = {}
            self.quantity_vars[tab] = {}
            self.non_optimized_vars[tab] = {}
            self.user_modified[tab] = {}  # 新增
            self.parts_qty_vars[tab] = {}
            for op in ops:
                name, std_t, opt_t, qty, unit, mtype = op
                self.check_vars[tab][name] = tk.BooleanVar(value=False)
                self.quantity_vars[tab][name] = tk.IntVar(value=qty)
                self.non_optimized_vars[tab][name] = tk.BooleanVar(value=False)
                self.parts_qty_vars[tab][name] = tk.IntVar(value=1)
                self.parts_qty_vars[tab][name].trace_add('write', lambda *args, t=tab: self.update_section_times(t))
                self.std_time_vars[(tab, name)] = tk.DoubleVar(value=std_t)
                self.original_std_times[(tab, name)] = std_t
                self.opt_time_dict[(tab, name)] = opt_t
                self.unit_dict[(tab, name)] = unit
                self.multi_type_dict[(tab, name)] = mtype

                # 绑定变化
                self.check_vars[tab][name].trace_add('write', lambda *args, t=tab: self.update_section_times(t))
                self.std_time_vars[(tab, name)].trace_add('write', lambda *args, t=tab: self.update_section_times(t))
                self.quantity_vars[tab][name].trace_add('write', lambda *args, t=tab, n=name: self.on_qty_change(t, n))
                self.non_optimized_vars[tab][name].trace_add('write', lambda *args, t=tab: self.update_section_times(t))

    def update_tuning_std_time(self, *args):
        tab = 'Tuning'
        rf_name = "RF TUNING1 estimated (if no other indication)"
        if rf_name in self.check_vars[tab]:
            tuning_type = self.tuning_type_var.get()
            if tuning_type in TUNING_TYPES:
                std_sec, opt_sec = TUNING_TYPES[tuning_type]
                self.std_time_vars[(tab, rf_name)].set(std_sec)  # Default to STANDARD for UI
            self.update_section_times(tab)

    def on_qty_change(self, tab, name, *args):
        """Qty 变化时标记用户修改"""
        if self.suppress_qty_modify:
            return
        if tab in ['Pre_Assembly', 'Assembly', 'Tuning', 'Testing']:
            self.user_modified[tab][name] = True
        self.update_section_times(tab)

    def get_housing(self):
        try:
            return int(self.housing_per_fg_var.get() or '2')
        except ValueError:
            return 2

    def get_n_cavities(self):
        try:
            return int(self.n_cavities_var.get() or '0')
        except ValueError:
            return 0

    def get_factor(self):
        cm = self.cm_var.get()
        product = self.product_var.get()
        if cm in FACTORS and product in FACTORS[cm]:
            return FACTORS[cm][product]
        return 1.0

    def on_housing_change(self, *args):
        """Housing per FG 变化时，更新默认 Qty"""
        housing = self.get_housing()
        section_ops = {
            'Pre_Assembly': PRE_ASSEMBLY_OPS,
            'Assembly': ASSEMBLY_OPS,
            'Tuning': TUNING_OPS,
            'Testing': TESTING_OPS
        }
        self.suppress_qty_modify = True
        for tab in ['Pre_Assembly', 'Assembly', 'Tuning', 'Testing']:
            ops = section_ops[tab]
            rf_name = "RF TUNING1 estimated (if no other indication)"
            for name in self.quantity_vars[tab]:
                if not self.user_modified.get(tab, {}).get(name, False):
                    default_qty = next(op[3] for op in ops if op[0] == name)
                    if tab == 'Tuning' and name == rf_name:
                        self.quantity_vars[tab][name].set(default_qty * housing)
                    else:
                        self.quantity_vars[tab][name].set(default_qty * housing)
        self.suppress_qty_modify = False
        self.update_all_times()

    def on_product_change(self, *args):
        prod = self.product_var.get()
        if prod in FPY:
            fpy_pct = f"{int(FPY[prod] * 100)}%"
            self.standard_fpy_var.set(fpy_pct)
            self.standard_custom_fpy_var.set(fpy_pct)
            self.optimized_fpy_var.set(fpy_pct)
            self.optimized_custom_fpy_var.set(fpy_pct)

    def on_standard_fpy_combo_change(self, *args):
        if self.standard_fpy_var.get() == 'Custom':
            self.standard_custom_fpy_entry.config(state='normal')
            # 关键：Custom 选中时，立即用输入框的值更新一次
            self.update_standard_fpy_rework()
        else:
            self.standard_custom_fpy_entry.config(state='disabled')
            self.standard_custom_fpy_var.set(self.standard_fpy_var.get())
            self.update_standard_fpy_rework()  # 也要更新！

    def on_optimized_fpy_combo_change(self, *args):
        if self.optimized_fpy_var.get() == 'Custom':
            self.optimized_custom_fpy_entry.config(state='normal')
            self.update_optimized_fpy_rework()
        else:
            self.optimized_custom_fpy_entry.config(state='disabled')
            self.optimized_custom_fpy_var.set(self.optimized_fpy_var.get())
            self.update_optimized_fpy_rework()

    def update_standard_fpy_rework(self, *args):
        try:
            if self.standard_fpy_var.get() == 'Custom':
                txt = self.standard_custom_fpy_var.get().strip('% ')
                if txt.isdigit():
                    fpy_pct = int(txt)
                else:
                    fpy_pct = 88
            else:
                fpy_pct = int(self.standard_fpy_var.get().strip('%'))
            
            # 限制合理范围
            fpy_pct = max(0, min(100, fpy_pct))
            rework_pct = 100 - fpy_pct
            self.standard_rework_var.set(f"{rework_pct}%")
        except:
            self.standard_rework_var.set("12%")

    def update_optimized_fpy_rework(self, *args):
        try:
            if self.optimized_fpy_var.get() == 'Custom':
                txt = self.optimized_custom_fpy_var.get().strip('% ')
                if txt.isdigit():
                    fpy_pct = int(txt)
                else:
                    fpy_pct = 88
            else:
                fpy_pct = int(self.optimized_fpy_var.get().strip('%'))
            
            fpy_pct = max(0, min(100, fpy_pct))
            rework_pct = 100 - fpy_pct
            self.optimized_rework_var.set(f"{rework_pct}%")
        except:
            self.optimized_rework_var.set("12%")

    def get_standard_fpy(self):
        try:
            if self.standard_fpy_var.get() == 'Custom':
                fpy_str = self.standard_custom_fpy_var.get()
            else:
                fpy_str = self.standard_fpy_var.get()
            return float(fpy_str.strip('%')) / 100
        except ValueError:
            return 0.88  # 默认值

    def get_optimized_fpy(self):
        try:
            if self.optimized_fpy_var.get() == 'Custom':
                fpy_str = self.optimized_custom_fpy_var.get()
            else:
                fpy_str = self.optimized_fpy_var.get()
            return float(fpy_str.strip('%')) / 100
        except ValueError:
            return 0.88  # 默认值

    def get_fpy_display(self, mode):
        if mode == 'STANDARD':
            if self.standard_fpy_var.get() == 'Custom':
                return self.standard_custom_fpy_var.get()
            else:
                return self.standard_fpy_var.get()
        else:
            if self.optimized_fpy_var.get() == 'Custom':
                return self.optimized_custom_fpy_var.get()
            else:
                return self.optimized_fpy_var.get()

    def setup_ui(self):
        # ==================== Basic Configuration ====================
        basic_frame = ttk.LabelFrame(self.main_frame, text=" Basic Configuration ", padding=12)
        basic_frame.pack(fill='x', pady=(0, 8))

        # 第一行：Product / CM / Tuning Type
        ttk.Label(basic_frame, text="Product:").grid(row=0, column=0, sticky='e', padx=(0,8))
        ttk.Combobox(basic_frame, textvariable=self.product_var, values=PRODUCTS, width=15).grid(row=0, column=1, sticky='w', padx=(0,30))

        ttk.Label(basic_frame, text="CM:").grid(row=0, column=2, sticky='e', padx=(20,8))
        ttk.Combobox(basic_frame, textvariable=self.cm_var, values=list(FACTORS.keys()), width=15).grid(row=0, column=3, sticky='w', padx=(0,30))

        ttk.Label(basic_frame, text="Tuning Type:").grid(row=0, column=4, sticky='e', padx=(20,8))
        ttk.Combobox(basic_frame, textvariable=self.tuning_type_var, values=list(TUNING_TYPES.keys()), width=15).grid(row=0, column=5, sticky='w')

        # ==================== Run Types ====================
        run_frame = ttk.LabelFrame(self.main_frame, text=" Run Types ", padding=12)
        run_frame.pack(fill='x', pady=(0, 12))

        # Modes
        ttk.Label(run_frame, text="Modes:").grid(row=0, column=0, sticky='w', padx=(0,8))
        modes_frame = ttk.Frame(run_frame)
        modes_frame.grid(row=0, column=1, columnspan=5, sticky='w')
        ttk.Checkbutton(modes_frame, text="STANDARD", variable=self.standard_mode_var).pack(side='left', padx=5)
        ttk.Checkbutton(modes_frame, text="OPTIMIZED", variable=self.optimized_mode_var).pack(side='left', padx=5)

        # STANDARD FPY and Rework
        ttk.Label(run_frame, text="STANDARD FPY:").grid(row=1, column=0, sticky='w', padx=(0,8), pady=(10,0))
        standard_fpy_frame = ttk.Frame(run_frame)
        standard_fpy_frame.grid(row=1, column=1, columnspan=2, sticky='w', pady=(10,0))
        self.standard_fpy_combo = ttk.Combobox(standard_fpy_frame, textvariable=self.standard_fpy_var,
                                               values=[f"{int(v*100)}%" for v in sorted(set(FPY.values()))] + ['Custom'],
                                               width=10, state='readonly')
        self.standard_fpy_combo.pack(side='left')
        self.standard_custom_fpy_entry = ttk.Entry(standard_fpy_frame, textvariable=self.standard_custom_fpy_var, width=8, state='disabled')
        self.standard_custom_fpy_entry.pack(side='left', padx=(5,0))

        ttk.Label(run_frame, text="STANDARD Rework:").grid(row=1, column=3, sticky='e', padx=(20,8), pady=(10,0))
        ttk.Label(run_frame, textvariable=self.standard_rework_var, width=15).grid(row=1, column=4, sticky='w', pady=(10,0))

        # OPTIMIZED FPY and Rework
        ttk.Label(run_frame, text="OPTIMIZED FPY:").grid(row=2, column=0, sticky='w', padx=(0,8), pady=(10,0))
        optimized_fpy_frame = ttk.Frame(run_frame)
        optimized_fpy_frame.grid(row=2, column=1, columnspan=2, sticky='w', pady=(10,0))
        self.optimized_fpy_combo = ttk.Combobox(optimized_fpy_frame, textvariable=self.optimized_fpy_var,
                                                values=[f"{int(v*100)}%" for v in sorted(set(FPY.values()))] + ['Custom'],
                                                width=10, state='readonly')
        self.optimized_fpy_combo.pack(side='left')
        self.optimized_custom_fpy_entry = ttk.Entry(optimized_fpy_frame, textvariable=self.optimized_custom_fpy_var, width=8, state='disabled')
        self.optimized_custom_fpy_entry.pack(side='left', padx=(5,0))

        ttk.Label(run_frame, text="OPTIMIZED Rework:").grid(row=2, column=3, sticky='e', padx=(20,8), pady=(10,0))
        ttk.Label(run_frame, textvariable=self.optimized_rework_var, width=15).grid(row=2, column=4, sticky='w', pady=(10,0))

        # ==================== Routing Parameters ====================
        param_frame = ttk.LabelFrame(self.main_frame, text=" Routing Parameters ", padding=12)
        param_frame.pack(fill='x', pady=(0, 12))

        ttk.Label(param_frame, text="N° Cavities:").grid(row=0, column=0, sticky='w', padx=(0,10))
        ttk.Entry(param_frame, textvariable=self.n_cavities_var, width=12, justify='center').grid(row=0, column=1, padx=(0,40))

        ttk.Label(param_frame, text="Housing per FG:").grid(row=0, column=2, sticky='w', padx=(20,10))
        ttk.Entry(param_frame, textvariable=self.housing_per_fg_var, width=12, justify='center').grid(row=0, column=3, padx=(0,40))

        # P/N in Routing Parameters
        ttk.Label(param_frame, text="P/N:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=(0,10), pady=(10,0))
        ttk.Entry(param_frame, textvariable=self.pn_var, width=28, font=('Arial', 10)).grid(row=1, column=1, columnspan=3, sticky='w', pady=(10,0))

        # Notebook
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(0,10))

        for tab_name, ops_list in [('Pre_Assembly', PRE_ASSEMBLY_OPS), ('Assembly', ASSEMBLY_OPS),
                              ('Tuning', TUNING_OPS), ('Testing', TESTING_OPS),
                              ('Quality Check', QUALITY_OPS), ('Packaging', PACKAGING_OPS)]:
            self.create_tab(tab_name, ops_list)

        # Buttons
        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.pack(pady=12, anchor='w')
        ttk.Button(btn_frame, text="Calculate Total Time", command=self.calculate_detailed, width=24).pack(side='left', padx=12)
        ttk.Button(btn_frame, text="Export to Excel", command=self.export_to_excel, style="Accent.TButton", width=24).pack(side='left', padx=12)
        ttk.Button(btn_frame, text="Reset All", command=self.reset_all, width=20).pack(side='left', padx=12)
        
        # Results
        self.results = scrolledtext.ScrolledText(self.main_frame, height=16, font=("Courier New", 10), bg='#f8f8f8')
        self.results.pack(fill='both', expand=True, padx=10, pady=(0,12))

        # Windows 滚轮绑定
        self.root.bind_all("<MouseWheel>", self.on_mousewheel)      

    ###############################################################
    # Windows 滚轮事件处理（同步滚动）
    ###############################################################
    def on_mousewheel(self, event):
        delta = int(-1 * (event.delta / 120))

        if self.results.winfo_exists():
            self.results.yview_scroll(delta, "units")

        current_tab = self.notebook.nametowidget(self.notebook.select())
        for child in current_tab.winfo_children():
            if isinstance(child, tk.Canvas):
                child.yview_scroll(delta, "units")

    ###############################################################
 
    
    def create_tab(self, tab_name, ops_list):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=f" {tab_name} ")
        self.calc_labels[tab_name] = {}

        canvas = tk.Canvas(tab)
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas)
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        has_parts_qty = True
        frame.grid_columnconfigure(6, weight=1)
        headers = ["Select", "Non-Optimized", "N° of P", "Qty", "Std Time", "Calc Time (min)", "Operation"]
        op_col = 6

        # 头部 (使用 tk.Label 以设置 bg)
        header_bg = '#d9d9d9'
        for col, text in enumerate(headers):
            # 注意：这里使用 tk.Label 替代了 ttk.Label
            lbl = tk.Label(frame, text=text, font=('Courier New', 9, 'bold'),
                             bg=header_bg, anchor='w', padx=4)
            lbl.grid(row=0, column=col, padx=3, pady=5, sticky='w')

        # 数据行
        for i, (name, std_t, opt_t, qty, unit, mtype) in enumerate(ops_list, 1):
            
            # 隔行变色逻辑: 仅 Operation 列使用此颜色
            row_color = "#f0f2f5" if i % 2 == 0 else "#ffffff" # 偶数行使用较浅的颜色，奇数行使用白色
            key = (tab_name, name)
            
            # Select (使用 ttk.Checkbutton 保持主题外观)
            ttk.Checkbutton(frame, variable=self.check_vars[tab_name][name]).grid(
                row=i, column=0, padx=6, pady=2, sticky='w')
            
            # Non-Optimized
            ttk.Checkbutton(frame, variable=self.non_optimized_vars[tab_name][name]).grid(
                row=i, column=1, padx=6, pady=2, sticky='w')
            
            col = 2
            # Parts Qty
            ttk.Entry(frame, textvariable=self.parts_qty_vars[tab_name][name],
                      width=6, justify='center').grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1
            
            # Qty (使用 ttk.Entry 保持主题外观)
            ttk.Entry(frame, textvariable=self.quantity_vars[tab_name][name],
                      width=6, justify='center').grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1

            # Std Time + unit (使用 ttk.Frame 容纳 ttk.Entry 和 ttk.Label)
            tf = ttk.Frame(frame)
            tf.grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1
            ttk.Entry(tf, textvariable=self.std_time_vars[key],
                      width=8, justify='center').pack(side='left')
            ttk.Label(tf, text=unit, width=4).pack(side='left')

            # Calc time (使用 tk.Label 以控制背景色为白色，与 Entry/Checkbutton 对齐)
            # 注意：这里使用 tk.Label 替代了 ttk.Label
            calc_lbl = tk.Label(frame, text="0.00",
                                font=('Courier New', 10),
                                fg='darkgreen',
                                bg='#ffffff') # 保持白色背景
            calc_lbl.grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1
            self.calc_labels[tab_name][name] = calc_lbl

            # Operation (使用 tk.Label 以设置隔行变色)
            # 注意：这里使用 tk.Label 替代了 ttk.Label
            op_lbl = tk.Label(frame, text=name, anchor='w', justify='left',
                              wraplength=680, bg=row_color) # 设置隔行颜色
            op_lbl.grid(row=i, column=op_col, sticky='ew', padx=6, pady=2) # sticky='ew' 使其填充 op_col
            
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def update_section_times(self, tab_name):
        # Use STANDARD for UI display
        mode = 'STANDARD'
        n_cav = self.get_n_cavities()
        factor = self.get_factor()
        ops_list = {'Pre_Assembly': PRE_ASSEMBLY_OPS, 'Assembly': ASSEMBLY_OPS, 'Tuning': TUNING_OPS,
                    'Testing': TESTING_OPS, 'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS}[tab_name]

        rf_name = "RF TUNING1 estimated (if no other indication)"
        check_name = "Tuning check after cleaning"
        rf_minutes = 0.0
        if tab_name == 'Tuning':
            if self.check_vars[tab_name][rf_name].get():
                std_t_rf = self.safe_get_double((tab_name, rf_name))
                opt_t_rf = self.opt_time_dict[(tab_name, rf_name)]
                t_per_rf = std_t_rf if mode == 'STANDARD' else opt_t_rf
                qty_rf = self.safe_get_int(self.quantity_vars[tab_name][rf_name])
                unit_rf = self.unit_dict[(tab_name, rf_name)]
                rf_minutes = (t_per_rf / 60.0 if unit_rf == 'sec' else t_per_rf) * qty_rf * n_cav
                self.std_time_vars[(tab_name, check_name)].set(round(0.1 * rf_minutes, 2))
            else:
                self.std_time_vars[(tab_name, check_name)].set(self.original_std_times[(tab_name, check_name)])

        for name, _, opt_t, _, unit, mtype in ops_list:
            if not self.check_vars[tab_name][name].get():
                self.calc_labels[tab_name][name].config(text="0.00")
                continue
            std_t = self.safe_get_double((tab_name, name))
            t_per = std_t if mode == 'STANDARD' else opt_t
            qty = self.safe_get_int(self.quantity_vars[tab_name][name])
            if tab_name == 'Tuning' and name == rf_name:
                minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty * n_cav
            else:
                minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty
            parts_qty = self.safe_get_int(self.parts_qty_vars[tab_name][name])
            minutes *= parts_qty
            self.calc_labels[tab_name][name].config(text=f"{minutes:8.2f}")

    def update_all_times(self, *args):
        for tab in self.check_vars:
            self.update_section_times(tab)

    def get_section_total(self, tab_name, ops_list, mode):
        total = 0.0
        n_cav = self.get_n_cavities()
        rf_name = "RF TUNING1 estimated (if no other indication)"
        check_name = "Tuning check after cleaning"
        rf_checked = self.check_vars[tab_name].get(rf_name, tk.BooleanVar(value=False)).get() if tab_name == 'Tuning' else False
        rf_base = 0.0
        for name, std_t_default, opt_t, _, unit, mtype in ops_list:
            if not self.check_vars[tab_name][name].get(): continue
            if mode == 'OPTIMIZED' and self.non_optimized_vars[tab_name][name].get(): continue
            if tab_name == 'Tuning' and name == rf_name:
                tuning_type = self.tuning_type_var.get()
                if tuning_type in TUNING_TYPES:
                    std_sec, opt_sec = TUNING_TYPES[tuning_type]
                    t_per = std_sec if mode == 'STANDARD' else opt_sec
                else:
                    t_per = std_t_default if mode == 'STANDARD' else opt_t
                qty = self.safe_get_int(self.quantity_vars[tab_name][name])
                base = (t_per / 60.0 if unit == 'sec' else t_per) * qty * n_cav
                rf_base = base
            elif tab_name == 'Tuning' and name == check_name:
                if rf_checked:
                    base = 0.1 * rf_base
                else:
                    t_per = std_t_default if mode == 'STANDARD' else opt_t
                    qty = self.safe_get_int(self.quantity_vars[tab_name][name])
                    base = (t_per if unit == 'min' else t_per / 60.0) * qty
            else:
                t_per = std_t_default if mode == 'STANDARD' else opt_t
                qty = self.safe_get_int(self.quantity_vars[tab_name][name])
                base = (t_per / 60.0 if unit == 'sec' else t_per) * qty
            parts_qty = self.safe_get_int(self.parts_qty_vars[tab_name][name])
            base *= parts_qty
            total += base
        return total

    def reset_all(self):
        self.n_cavities_var.set('0')
        self.housing_per_fg_var.set('2')
        self.standard_mode_var.set(True)
        self.optimized_mode_var.set(False)
        self.pn_var.set("")
        self.standard_fpy_var.set('88%')
        self.standard_custom_fpy_var.set('88%')
        self.standard_rework_var.set('12%')
        self.optimized_fpy_var.set('88%')
        self.optimized_custom_fpy_var.set('88%')
        self.optimized_rework_var.set('12%')
        housing = self.get_housing()
        section_ops = {'Pre_Assembly': PRE_ASSEMBLY_OPS, 'Assembly': ASSEMBLY_OPS, 'Tuning': TUNING_OPS,
                       'Testing': TESTING_OPS, 'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS}
        self.suppress_qty_modify = True
        rf_name = "RF TUNING1 estimated (if no other indication)"
        check_name = "Tuning check after cleaning"
        for tab in self.check_vars:
            for var in self.check_vars[tab].values(): var.set(False)
            for var in self.non_optimized_vars[tab].values(): var.set(False)
            self.user_modified[tab] = {}
            for var in self.parts_qty_vars[tab].values():
                var.set(1)
            for name, var in self.quantity_vars[tab].items():
                for op in section_ops[tab]:
                    if op[0] == name:
                        default_qty = op[3]
                        if tab == 'Tuning' and name == check_name:
                            var.set(default_qty)
                        elif tab in ['Pre_Assembly', 'Assembly', 'Tuning', 'Testing'] or (tab == 'Tuning' and name == rf_name):
                            var.set(default_qty * housing)
                        else:
                            var.set(default_qty)
                        break
            for key in [k for k in self.std_time_vars if k[0] == tab]:
                self.std_time_vars[key].set(self.original_std_times[key])
        self.suppress_qty_modify = False
        self.update_all_times()
        self.results.delete(1.0, tk.END)
        self.standard_custom_fpy_entry.config(state='disabled')
        self.optimized_custom_fpy_entry.config(state='disabled')

    def calculate_detailed(self):
        self.update_all_times()
        product = self.product_var.get()
        cm = self.cm_var.get()
        n_cav = self.get_n_cavities()
        housing = self.get_housing()
        modules = housing
        tuning_type = self.tuning_type_var.get()
        pn = self.pn_var.get() or "N/A"

        factor = self.get_factor()
        selected_modes = []
        if self.standard_mode_var.get():
            selected_modes.append(('STANDARD', self.get_standard_fpy()))
        if self.optimized_mode_var.get():
            selected_modes.append(('OPTIMIZED', self.get_optimized_fpy()))

        if not selected_modes:
            messagebox.showwarning("Warning", "Please select at least one mode.")
            return

        result = ""
        for mode, fpy in selected_modes:
            pre_assembly_base = self.get_section_total('Pre_Assembly', PRE_ASSEMBLY_OPS, mode)
            #pre_assembly = pre_assembly_base * factor
            assembly_base = self.get_section_total('Assembly', ASSEMBLY_OPS, mode)
            #assembly = assembly_base * factor
            tuning_base = self.get_section_total('Tuning', TUNING_OPS, mode)
            #tuning_adjusted = tuning_base * factor
            testing_base = self.get_section_total('Testing', TESTING_OPS, mode)
            #testing = testing_base * factor
            quality = self.get_section_total('Quality Check', QUALITY_OPS, mode)
            packaging = self.get_section_total('Packaging', PACKAGING_OPS, mode)

            total_wo = pre_assembly_base + assembly_base + tuning_base + testing_base + quality + packaging
            rework_rate = 1 - fpy
            rework = (total_wo - pre_assembly_base - quality - packaging) * rework_rate
            total = total_wo + rework

            result += f""" ROUTING CALCULATION RESULT ({mode})
{'═' * 100}
P/N            : {pn}
Product        : {product:<12} CM: {cm:<12} FPY: {self.get_fpy_display(mode)}
Tuning Type    : {tuning_type}
Cavities × Modules : {n_cav} × {modules:<6} Housing/FG: {housing}
Mode           : {mode}

Pre_Assembly   : {pre_assembly_base:6.2f}
Assembly       : {assembly_base:6.2f}
Tuning         : {tuning_base:6.2f}
Testing        : {testing_base:6.2f}
Cleaning       : {0.0:6.2f}  (included in Assembly)
Quality Check  : {quality:6.2f}
Packaging      : {packaging:6.2f}
{'─' * 50}
Rework         : {rework:6.2f}
Total w/o Rework : {total_wo:6.2f}
╔{'═'*42}╗
║  Total Routing Time : {total:7.2f} minutes      ║
╚{'═'*42}╝

"""

        self.results.delete(1.0, tk.END)
        self.results.insert(tk.END, result)

    def export_to_excel(self):
        self.update_all_times()
        try:
            pn_raw = self.pn_var.get().strip()
            pn = "".join(c if c.isalnum() or c in "._-/\\" else "_" for c in pn_raw) if pn_raw else "NoPN"
            tuning = self.tuning_type_var.get().replace(" ", "_").replace("(", "").replace(")", "")
            suggested_name = f"Routing_{pn}_{tuning}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                title="保存 Routing 结果",
                initialfile=suggested_name
            )
            if not filename:
                return

            selected_modes = []
            if self.standard_mode_var.get():
                selected_modes.append(('STANDARD', self.get_standard_fpy(), self.get_fpy_display('STANDARD')))
            if self.optimized_mode_var.get():
                selected_modes.append(('OPTIMIZED', self.get_optimized_fpy(), self.get_fpy_display('OPTIMIZED')))

            if not selected_modes:
                messagebox.showwarning("Warning", "Please select at least one mode.")
                return

            factor = self.get_factor()
            n_cav = self.get_n_cavities()
            housing = self.get_housing()
            tuning_type = self.tuning_type_var.get()

            rows = []
            for tab, ops in [('Pre_Assembly', PRE_ASSEMBLY_OPS), ('Assembly', ASSEMBLY_OPS), ('Tuning', TUNING_OPS),
                            ('Testing', TESTING_OPS), ('Quality Check', QUALITY_OPS), ('Packaging', PACKAGING_OPS)]:
                for name, std_t, opt_t, _, unit, _ in ops:
                    if self.check_vars[tab][name].get():
                        parts_qty_val = self.safe_get_int(self.parts_qty_vars[tab][name])
                        qty = self.safe_get_int(self.quantity_vars[tab][name])
                        # STANDARD Calc Time
                        t_per_std = std_t
                        if tab == 'Tuning' and name == "RF TUNING1 estimated (if no other indication)":
                            if tuning_type in TUNING_TYPES:
                                std_sec, _ = TUNING_TYPES[tuning_type]
                                t_per_std = std_sec
                            minutes_std = (t_per_std / 60.0 if unit == 'sec' else t_per_std) * qty * n_cav
                        else:
                            minutes_std = (t_per_std / 60.0 if unit == 'sec' else t_per_std) * qty
                        minutes_std *= parts_qty_val

                        # OPTIMIZED Calc Time
                        if self.non_optimized_vars[tab][name].get():
                            minutes_opt = 0.0
                        else:
                            t_per_opt = opt_t
                            if tab == 'Tuning' and name == "RF TUNING1 estimated (if no other indication)":
                                if tuning_type in TUNING_TYPES:
                                    _, opt_sec = TUNING_TYPES[tuning_type]
                                    t_per_opt = opt_sec
                                minutes_opt = (t_per_opt / 60.0 if unit == 'sec' else t_per_opt) * qty * n_cav
                            else:
                                minutes_opt = (t_per_opt / 60.0 if unit == 'sec' else t_per_opt) * qty
                            minutes_opt *= parts_qty_val

                        # Std Time STANDARD and OPTIMIZED
                        if tab == 'Tuning' and name == "RF TUNING1 estimated (if no other indication)":
                            if tuning_type in TUNING_TYPES:
                                std_sec, opt_sec = TUNING_TYPES[tuning_type]
                                std_time_std = f"{std_sec:.2f} sec"
                                std_time_opt = f"{opt_sec:.2f} sec"
                            else:
                                std_time_std = f"{std_t:.2f} {unit}"
                                std_time_opt = f"{opt_t:.2f} {unit}"
                        else:
                            std_time_std = f"{std_t:.2f} {unit}"
                            std_time_opt = f"{opt_t:.2f} {unit}"

                        rows.append({
                            "Section": tab,
                            "Operation": name,
                            "Non-Optimized": self.non_optimized_vars[tab][name].get(),
                            "N° of P": parts_qty_val,
                            "Qty": qty,
                            "Std Time STANDARD": std_time_std,
                            "Std Time OPTIMIZED": std_time_opt,
                            "Calc Time STANDARD (min)": round(minutes_std, 3),
                            "Calc Time OPTIMIZED (min)": round(minutes_opt, 3)
                        })
            df_ops = pd.DataFrame(rows) if rows else pd.DataFrame({"Info": ["No operations selected"]})

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Configuration (common)
                config_data = [
                    ["P/N", pn_raw or "N/A"],
                    ["Product", self.product_var.get()],
                    ["CM", self.cm_var.get()],
                    ["Tuning Type", self.tuning_type_var.get()],
                    ["N° Cavities", n_cav],
                    ["Housing per FG", housing],
                    ["Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                ]
                pd.DataFrame(config_data, columns=["Parameter", "Value"]).to_excel(writer, sheet_name="Configuration", index=False)

                df_ops.to_excel(writer, sheet_name="Selected Operations", index=False)

                # Collect data for all modes
                items = ["Pre_Assembly", "Assembly", "Tuning", "Testing", "Quality Check", "Packaging", "Rework", "Total w/o Rework", "Total Routing Time"]
                all_values = []
                mode_names = []
                for mode, fpy, fpy_display in selected_modes:
                    pre_assembly_base = self.get_section_total('Pre_Assembly', PRE_ASSEMBLY_OPS, mode)
                    assembly_base = self.get_section_total('Assembly', ASSEMBLY_OPS, mode)
                    tuning_base = self.get_section_total('Tuning', TUNING_OPS, mode)
                    testing_base = self.get_section_total('Testing', TESTING_OPS, mode)
                    quality = self.get_section_total('Quality Check', QUALITY_OPS, mode)
                    packaging = self.get_section_total('Packaging', PACKAGING_OPS, mode)

                    total_wo_rework = pre_assembly_base + assembly_base + tuning_base + testing_base + quality + packaging
                    rework_rate = 1 - fpy
                    rework = (total_wo_rework - pre_assembly_base - quality - packaging) * rework_rate
                    final_total = total_wo_rework + rework

                    values = [
                        round(pre_assembly_base,3),
                        round(assembly_base,3),
                        round(tuning_base,3),
                        round(testing_base,3),
                        round(quality,3),
                        round(packaging,3),
                        round(rework,3),
                        round(total_wo_rework,3),
                        round(final_total,3)
                    ]
                    all_values.append(values)
                    mode_names.append(f"{mode} Base Time")

                summary_df = pd.DataFrame(all_values, columns=items)
                summary_df.insert(0, "Mode", mode_names)
                summary_df.to_excel(writer, sheet_name="Time Summary", index=False, startrow=1)

                # 美化
                wb = writer.book
                for ws in wb.worksheets:
                    if ws.title == "Time Summary":
                        ws['A1'] = f"P/N: {pn_raw or 'N/A'}"
                        ws['A1'].font = Font(bold=True)
                        ws['A1'].fill = PatternFill(start_color="D6EBFF", end_color="D6EBFF", fill_type="solid")
                        ws['A1'].alignment = Alignment(horizontal="left")
                        header_row = 2
                    else:
                        header_row = 1

                    for cell in ws[header_row]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D6EBFF", end_color="D6EBFF", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

                    for col in ws.columns:
                        max_len = 0
                        column = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[column].width = min(max_len + 4, 60)

                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    min_border_row = 1 if ws.title == "Time Summary" else header_row
                    for row in ws.iter_rows(min_row=min_border_row):
                        for cell in row:
                            cell.border = border

                # Special alignment for Configuration sheet
                config_ws = wb["Configuration"]
                for row in config_ws.iter_rows(min_row=2, max_row=config_ws.max_row, min_col=1, max_col=2):
                    param = row[0].value
                    if param in ["N° Cavities", "Housing per FG"]:
                        row[1].alignment = Alignment(horizontal='left')

            messagebox.showinfo("Success", f"Excel has been successfully exported!\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    set_app_icon(root) # 新增: 在创建 root 后设置图标        
    app = RoutingApp(root)
    root.mainloop()