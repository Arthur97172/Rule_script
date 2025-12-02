# -*- coding: utf-8 -*-
"""
Routing Estimation Tool - 工业级终极完整版
功能：
1. 新增 P/N 输入框
2. 导出文件名自动包含 P/N + Tuning Type
3. 一键导出美观 Excel（3个Sheet）
4. 完整工序数据 + 实时计算 + Reset All
5. Qty 默认根据 Housing per FG 自动更新（用户修改后保持不变）
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import scrolledtext
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def set_app_icon(window):
    """统一设置窗口图标，自动处理打包路径（静默失败，不输出任何信息）"""
    try:
        icon_path = resource_path("app.ico")
        window.iconbitmap(icon_path)
    except:
        pass  # 图标缺失时静默跳过，不影响程序运行


# ================================= 数据定义 =================================

PRODUCTS = ['TMA', 'DPX', 'TPX', 'PPX', 'QPX', 'IMF', 'LLC']

FACTORS = {
    'TIGER': {
        'TMA': 1.0, 'DPX': 1.1, 'TPX': 1.28, 'PPX': 0.85, 'QPX': 0.95,
        'IMF': 0.83, 'LLC': 0.69
    },
    'TATFOOK': {
        'TMA': 1.39, 'DPX': 1.39, 'TPX': 1.43, 'PPX': 1.04, 'QPX': 0.92,
        'IMF': 1.02, 'LLC': 0.84
    },
    'LUXSHARE': {
        'TMA': 1.24, 'DPX': 1.35, 'TPX': 1.68, 'PPX': 1.18, 'QPX': 0.84,
        'IMF': 1.04, 'LLC': 0.87
    }
}

FPY = {'DPX': 0.9, 'TPX': 0.88, 'QPX': 0.85, 'TMA': 0.9, 'IMF': 0.85, 'LLC': 0.85, 'PPX': 0.85}

PRE_ASSEMBLY_OPS = [
    ("Tuning screw glued to tuners or plug", 12, 12, 1, 'sec', 'housing'),
    ("Lid cover+tuning screw+nut+washer", 8, 8, 1, 'sec', 'housing'),
    ("Connector+conductor soldered", 45, 45, 1, 'sec', 'housing'),
    ("Connector+conductor screwed", 5, 5, 1, 'sec', 'housing'),
]

ASSEMBLY_OPS = [
    ("Low pass(or bias tee or coils) + thermalshrinking", 120, 120, 1, 'sec', 'housing'),
    ("Bias coil group", 30, 30, 1, 'sec', 'housing'),
    ("Lid cover + pin", 30, 30, 1, 'sec', 'housing'),
    ("Lid cover + line screwed", 5, 5, 1, 'sec', 'housing'),
    ("Lid cover + line soldered", 67, 67, 1, 'sec', 'housing'),
    ("Lid cover + connector", 15, 15, 1, 'sec', 'housing'),
    ("Lid cover + connector press in", 60, 60, 1, 'sec', 'housing'),
    ("Label for s/n", 5, 5, 1, 'sec', 'housing'),
    ("Housing + pin", 10, 10, 1, 'sec', 'housing'),
    ("Housing + connector flanged", 45, 45, 1, 'sec', 'housing'),
    ("Housing + connector screwed", 30, 30, 1, 'sec', 'housing'),
    ("Housing + junction screwed (time x 2 screws)", 10, 10, 1, 'sec', 'housing'),
    ("Housing + junction soldered by SOLDER STATION", 30, 30, 1, 'sec', 'housing'),
    ("Housing + junction soldered INDUCTION", 55, 55, 1, 'sec', 'housing'),
    ("Housing + line soldered by SOLDER STATION (complex)", 150, 150, 1, 'sec', 'housing'),
    ("Housing + line soldered INDUCTION (complex)", 150, 150, 1, 'sec', 'housing'),
    ("Housing + low pass(or bias tee)", 10, 10, 1, 'sec', 'housing'),
    ("Ret connector", 20, 20, 1, 'sec', 'housing'),
    ("Resonator", 7, 7, 1, 'sec', 'housing'),
    ("Housing + probe (including support fixing)", 9, 9, 1, 'sec', 'housing'),
    ("Apply solder paste for LID SOLDERING process", 300, 300, 1, 'sec', 'housing'),
    ("Tuning lid cover screwing", 3, 3, 1, 'sec', 'housing'),
    ("PWBA installation (including soldering)", 8, 8, 1, 'sec', 'housing'),
    ("Cable sub-assembly(or LED)", 15, 15, 1, 'sec', 'housing'),
    ("Outer cover (time x screw)", 4, 4, 1, 'sec', 'housing'),
    ("PWBA+Nickel Silver Cover placing", 3, 3, 1, 'sec', 'housing'),
    ("IP outer label ASSEMBLY", 60, 60, 1, 'sec', 'housing'),
    ("Housing + venting membrane", 10, 10, 1, 'sec', 'housing'),
    ("Housing + bracket +screw + nut + washer", 120, 120, 1, 'sec', 'housing'),
    ("Tuning lid open and close (time x screw)", 4, 4, 1, 'sec', 'housing'),
    ("Cleaning", 4.0, 4.0, 1, 'min', 'housing'),
]

TUNING_OPS = [
    ("RF TUNING1 estimated (if no other indication)", 108, 108, 1, 'sec', 'tuning'),  
    ("TUNING 2 (E35)", 8.0, 8.0, 1, 'min', 'tuning'),
    ("Tuning check after cleaning", 5.0, 5.0, 1, 'min', 'tuning'),
]

TESTING_OPS = [
    ("PCBA programming (FW uploading)", 1.0, 1.0, 1, 'min', 'none'),
    ("NOISE FIGURE(mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("IP3(optional for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("AC/DC or Generic Measures", 3.0, 3.0, 1, 'min', 'none'),
    ("IMD (STANDARD)", 1.5, 0.0, 1, 'min', 'none'),
    ("IMD (OPTIMIZED)", 0.0, 1.5, 1, 'min', 'none'),
    ("LINEARIZATION (mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("PID(mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("TDR on TMS (FINAL TEST)", 3.0, 3.0, 1, 'min', 'none'),
    ("PRESSURE TEST", 2.0, 2.0, 1, 'min', 'none'),
    ("LED", 2.0, 2.0, 1, 'min', 'none'),
]

QUALITY_OPS = [("Final check 100% (VISUAL INSPECTION)", 1.0, 1.0, 1, 'min', 'none')]
PACKAGING_OPS = [("Packaging", 5.2, 5.2, 1, 'min', 'none')]

TUNING_TYPES = {
    'DPX': (80, 72),
    'TPX': (100, 90),
    'PPX': (100, 90),
    'QPX': (100, 90),
    'TMA': (100, 90),
    'LLC': (120, 108),
    'IMF': (120, 108),
    'Demanding perf': (120, 108),
}

# ==============================================================================

class RoutingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Routing Estimation Tool - V1.0")
        self.root.geometry("1080x900")
        self.root.minsize(900, 720)

        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill='both', expand=True, padx=15, pady=12)

        # 新增 P/N
        self.pn_var = tk.StringVar(value="")

        self.product_var = tk.StringVar(value='TPX')
        self.cm_var = tk.StringVar(value='TIGER')
        self.mode_var = tk.StringVar(value='STANDARD')
        self.fpy_var = tk.StringVar(value='88%')
        self.custom_fpy_var = tk.StringVar(value='88%')
        self.rework_var = tk.StringVar(value='12%')
        self.n_cavities_var = tk.StringVar(value='0')
        self.housing_per_fg_var = tk.StringVar(value='2')
        self.tuning_type_var = tk.StringVar(value='DPX')

        self.suppress_qty_modify = False

        self.original_std_times = {}
        self.check_vars = {}
        self.quantity_vars = {}
        self.parts_qty_vars = {}
        self.std_time_vars = {}
        self.opt_time_dict = {}
        self.unit_dict = {}
        self.multi_type_dict = {}
        self.calc_labels = {}

        # 新增：用户修改标记（用于 Qty 自动更新）
        self.user_modified = {}

        self.setup_data()
        self.setup_ui()

        self.mode_var.trace_add('write', self.update_all_times)
        self.mode_var.trace_add('write', self.update_tuning_std_time)
        self.housing_per_fg_var.trace_add('write', self.on_housing_change)
        self.n_cavities_var.trace_add('write', self.update_all_times)
        self.product_var.trace_add('write', self.on_product_change)
        self.fpy_var.trace_add('write', self.on_fpy_combo_change)
        self.custom_fpy_var.trace_add('write', self.update_fpy_rework)
        self.tuning_type_var.trace_add('write', self.update_tuning_std_time)
        self.cm_var.trace_add('write', self.update_all_times)

        self.update_all_times()
        self.on_product_change()
        self.update_fpy_rework()
        self.on_housing_change()
        self.update_tuning_std_time()

    def setup_data(self):
        sections = {
            'Pre_Assembly': PRE_ASSEMBLY_OPS,
            'Assembly': ASSEMBLY_OPS,
            'Tuning': TUNING_OPS,
            'Testing': TESTING_OPS,
            'Quality Check': QUALITY_OPS,
            'Packaging': PACKAGING_OPS
        }
        for tab, ops in sections.items():
            self.check_vars[tab] = {}
            self.quantity_vars[tab] = {}
            self.user_modified[tab] = {}  # 新增
            if tab in ['Pre_Assembly', 'Assembly']:
                self.parts_qty_vars[tab] = {}
            for op in ops:
                name, std_t, opt_t, qty, unit, mtype = op
                self.check_vars[tab][name] = tk.BooleanVar(value=False)
                self.quantity_vars[tab][name] = tk.IntVar(value=qty)
                if tab in ['Pre_Assembly', 'Assembly']:
                    self.parts_qty_vars[tab][name] = tk.IntVar(value=1)
                    self.parts_qty_vars[tab][name].trace_add('write', lambda *args, t=tab: self.update_section_times(t))
                self.std_time_vars[(tab, name)] = tk.DoubleVar(value=std_t)
                self.original_std_times[(tab, name)] = std_t
                self.opt_time_dict[(tab, name)] = opt_t
                self.unit_dict[(tab, name)] = unit
                self.multi_type_dict[(tab, name)] = mtype

                # 绑定变化
                self.check_vars[tab][name].trace_add('write', lambda *args, t=tab: self.update_section_times(t))
                self.std_time_vars[(tab, name)].trace_add('write', lambda *args, t=tab: self.update_section_times(t))
                self.quantity_vars[tab][name].trace_add('write', lambda *args, t=tab, n=name: self.on_qty_change(t, n))

    def update_tuning_std_time(self, *args):
        tab = 'Tuning'
        rf_name = "RF TUNING1 estimated (if no other indication)"
        if rf_name in self.check_vars[tab]:
            tuning_type = self.tuning_type_var.get()
            mode = self.mode_var.get()
            if tuning_type in TUNING_TYPES:
                std_sec, opt_sec = TUNING_TYPES[tuning_type]
                if mode == 'STANDARD':
                    self.std_time_vars[(tab, rf_name)].set(std_sec)
                else:
                    self.std_time_vars[(tab, rf_name)].set(opt_sec)
            self.update_section_times(tab)

    def on_qty_change(self, tab, name, *args):
        """Qty 变化时标记用户修改"""
        if self.suppress_qty_modify:
            return
        if tab in ['Pre_Assembly', 'Assembly', 'Tuning']:
            self.user_modified[tab][name] = True
        self.update_section_times(tab)

    def get_housing(self):
        try:
            return int(self.housing_per_fg_var.get() or '2')
        except ValueError:
            return 2

    def get_n_cavities(self):
        try:
            return int(self.n_cavities_var.get() or '0')
        except ValueError:
            return 0

    def get_factor(self):
        cm = self.cm_var.get()
        product = self.product_var.get()
        if cm in FACTORS and product in FACTORS[cm]:
            return FACTORS[cm][product]
        return 1.0

    def on_housing_change(self, *args):
        """Housing per FG 变化时，更新默认 Qty"""
        housing = self.get_housing()
        section_ops = {
            'Pre_Assembly': PRE_ASSEMBLY_OPS,
            'Assembly': ASSEMBLY_OPS,
            'Tuning': TUNING_OPS
        }
        self.suppress_qty_modify = True
        for tab in ['Pre_Assembly', 'Assembly', 'Tuning']:
            ops = section_ops[tab]
            rf_name = "RF TUNING1 estimated (if no other indication)"
            for name in self.quantity_vars[tab]:
                if not self.user_modified.get(tab, {}).get(name, False):
                    default_qty = next(op[3] for op in ops if op[0] == name)
                    if tab == 'Tuning' and name == rf_name:
                        self.quantity_vars[tab][name].set(default_qty * housing)
                    else:
                        self.quantity_vars[tab][name].set(default_qty * housing)
        self.suppress_qty_modify = False
        self.update_all_times()

    def on_product_change(self, *args):
        prod = self.product_var.get()
        if prod in FPY:
            fpy_pct = f"{int(FPY[prod] * 100)}%"
            self.fpy_var.set(fpy_pct)
            self.custom_fpy_var.set(fpy_pct)

    def on_fpy_combo_change(self, *args):
        if self.fpy_var.get() == 'Custom':
            self.custom_fpy_entry.config(state='normal')
        else:
            self.custom_fpy_entry.config(state='disabled')
            self.custom_fpy_var.set(self.fpy_var.get())
        self.update_fpy_rework()

    def update_fpy_rework(self, *args):
        try:
            if self.fpy_var.get() == 'Custom':
                fpy_str = self.custom_fpy_var.get()
            else:
                fpy_str = self.fpy_var.get()
            fpy_num = float(fpy_str.strip('%')) / 100
            rework_pct = int((1 - fpy_num) * 100)
            self.rework_var.set(f"{rework_pct}%")
        except ValueError:
            pass

    def get_fpy(self):
        try:
            if self.fpy_var.get() == 'Custom':
                fpy_str = self.custom_fpy_var.get()
            else:
                fpy_str = self.fpy_var.get()
            return float(fpy_str.strip('%')) / 100
        except ValueError:
            return 0.88  # 默认值

    def get_fpy_display(self):
        if self.fpy_var.get() == 'Custom':
            return self.custom_fpy_var.get()
        else:
            return self.fpy_var.get()

    def setup_ui(self):
        # Basic Configuration
        basic_frame = ttk.LabelFrame(self.main_frame, text=" Basic Configuration ", padding=12)
        basic_frame.pack(fill='x', pady=(0, 8))

        # 第一行：Product / CM / Tuning Type
        ttk.Label(basic_frame, text="Product:").grid(row=0, column=0, sticky='e', padx=(0,8))
        ttk.Combobox(basic_frame, textvariable=self.product_var, values=PRODUCTS, width=15).grid(row=0, column=1, sticky='w', padx=(0,30))

        ttk.Label(basic_frame, text="CM:").grid(row=0, column=2, sticky='e', padx=(20,8))
        ttk.Combobox(basic_frame, textvariable=self.cm_var, values=list(FACTORS.keys()), width=15).grid(row=0, column=3, sticky='w', padx=(0,30))

        ttk.Label(basic_frame, text="Tuning Type:").grid(row=0, column=4, sticky='e', padx=(20,8))
        ttk.Combobox(basic_frame, textvariable=self.tuning_type_var, values=list(TUNING_TYPES.keys()), width=15).grid(row=0, column=5, sticky='w')

        # 第二行：Mode / FPY / Rework
        ttk.Label(basic_frame, text="Mode:").grid(row=1, column=0, sticky='e', padx=(0,8), pady=(10,0))
        ttk.Combobox(basic_frame, textvariable=self.mode_var, values=['STANDARD', 'OPTIMIZED'], width=15).grid(row=1, column=1, sticky='w', padx=(0,30), pady=(10,0))

        ttk.Label(basic_frame, text="FPY:").grid(row=1, column=2, sticky='e', padx=(20,8), pady=(10,0))

        # FPY Frame for Combo and Custom Entry
        fpy_frame = ttk.Frame(basic_frame)
        fpy_frame.grid(row=1, column=3, sticky='w', pady=(10,0))
        self.fpy_combo = ttk.Combobox(fpy_frame, textvariable=self.fpy_var, 
                                      values=[f"{int(v*100)}%" for v in sorted(set(FPY.values()))] + ['Custom'], 
                                      width=10, state='readonly')
        self.fpy_combo.pack(side='left')
        self.custom_fpy_entry = ttk.Entry(fpy_frame, textvariable=self.custom_fpy_var, width=8, state='disabled')
        self.custom_fpy_entry.pack(side='left', padx=(5,0))

        ttk.Label(basic_frame, text="Rework:").grid(row=1, column=4, sticky='e', padx=(20,8), pady=(10,0))
        ttk.Label(basic_frame, textvariable=self.rework_var, width=15).grid(row=1, column=5, sticky='w', pady=(10,0))

        # Routing Parameters
        param_frame = ttk.LabelFrame(self.main_frame, text=" Routing Parameters ", padding=12)
        param_frame.pack(fill='x', pady=(0, 12))

        ttk.Label(param_frame, text="N° Cavities:").grid(row=0, column=0, sticky='w', padx=(0,10))
        ttk.Entry(param_frame, textvariable=self.n_cavities_var, width=12, justify='center').grid(row=0, column=1, padx=(0,40))

        ttk.Label(param_frame, text="Housing per FG:").grid(row=0, column=2, sticky='w', padx=(20,10))
        ttk.Entry(param_frame, textvariable=self.housing_per_fg_var, width=12, justify='center').grid(row=0, column=3, padx=(0,40))

        # P/N in Routing Parameters
        ttk.Label(param_frame, text="P/N:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=(0,10), pady=(10,0))
        ttk.Entry(param_frame, textvariable=self.pn_var, width=28, font=('Arial', 10)).grid(row=1, column=1, columnspan=3, sticky='w', pady=(10,0))

        # Notebook
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(0,10))

        for tab_name, ops in [('Pre_Assembly', PRE_ASSEMBLY_OPS), ('Assembly', ASSEMBLY_OPS),
                              ('Tuning', TUNING_OPS), ('Testing', TESTING_OPS),
                              ('Quality Check', QUALITY_OPS), ('Packaging', PACKAGING_OPS)]:
            self.create_tab(tab_name, ops)

        # Buttons
        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.pack(pady=12)
        ttk.Button(btn_frame, text="Calculate Total Time", command=self.calculate_detailed, width=24).pack(side='left', padx=12)
        ttk.Button(btn_frame, text="Reset All", command=self.reset_all, width=20).pack(side='left', padx=12)
        ttk.Button(btn_frame, text="Export to Excel", command=self.export_to_excel, style="Accent.TButton", width=24).pack(side='left', padx=12)

        # Results
        self.results = scrolledtext.ScrolledText(self.main_frame, height=16, font=("Courier New", 10), bg='#f8f8f8')
        self.results.pack(fill='both', expand=True, padx=10, pady=(0,12))

        # Windows 滚轮绑定
        self.root.bind_all("<MouseWheel>", self.on_mousewheel)      


    ###############################################################
    # Windows 滚轮事件处理（同步滚动）
    ###############################################################
    def on_mousewheel(self, event):
        delta = int(-1 * (event.delta / 120))

        if self.results.winfo_exists():
            self.results.yview_scroll(delta, "units")

        current_tab = self.notebook.nametowidget(self.notebook.select())
        for child in current_tab.winfo_children():
            if isinstance(child, tk.Canvas):
                child.yview_scroll(delta, "units")

    ###############################################################

    
    def create_tab(self, tab_name, ops_list):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=f" {tab_name} ")
        self.calc_labels[tab_name] = {}

        canvas = tk.Canvas(tab)
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas)
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        has_parts_qty = tab_name in ['Pre_Assembly', 'Assembly']
        if has_parts_qty:
            frame.grid_columnconfigure(5, weight=1)
            headers = ["Select", "Parts Qty", "Qty", "Std Time", "Calc Time (min)", "Operation"]
            op_col = 5
        else:
            frame.grid_columnconfigure(4, weight=1)
            headers = ["Select", "Qty", "Std Time", "Calc Time (min)", "Operation"]
            op_col = 4

        # 头部 (使用 tk.Label 以设置 bg)
        header_bg = '#d9d9d9'
        for col, text in enumerate(headers):
            # 注意：这里使用 tk.Label 替代了 ttk.Label
            lbl = tk.Label(frame, text=text, font=('Courier New', 9, 'bold'),
                             bg=header_bg, anchor='w', padx=4)
            lbl.grid(row=0, column=col, padx=3, pady=5, sticky='w')

        # 数据行
        for i, (name, std_t, opt_t, qty, unit, mtype) in enumerate(ops_list, 1):
            
            # 隔行变色逻辑: 仅 Operation 列使用此颜色
            row_color = "#f0f2f5" if i % 2 == 0 else "#ffffff" # 偶数行使用较浅的颜色，奇数行使用白色
            key = (tab_name, name)
            
            # Select (使用 ttk.Checkbutton 保持主题外观)
            ttk.Checkbutton(frame, variable=self.check_vars[tab_name][name]).grid(
                row=i, column=0, padx=6, pady=2, sticky='w')
            
            col = 1
            if has_parts_qty:
                # Parts Qty
                ttk.Entry(frame, textvariable=self.parts_qty_vars[tab_name][name],
                          width=6, justify='center').grid(row=i, column=col, padx=3, pady=2, sticky='w')
                col += 1
            
            # Qty (使用 ttk.Entry 保持主题外观)
            ttk.Entry(frame, textvariable=self.quantity_vars[tab_name][name],
                      width=6, justify='center').grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1

            # Std Time + unit (使用 ttk.Frame 容纳 ttk.Entry 和 ttk.Label)
            tf = ttk.Frame(frame)
            tf.grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1
            ttk.Entry(tf, textvariable=self.std_time_vars[key],
                      width=8, justify='center').pack(side='left')
            ttk.Label(tf, text=unit, width=4).pack(side='left')

            # Calc time (使用 tk.Label 以控制背景色为白色，与 Entry/Checkbutton 对齐)
            # 注意：这里使用 tk.Label 替代了 ttk.Label
            calc_lbl = tk.Label(frame, text="0.00",
                                font=('Courier New', 10),
                                fg='darkgreen',
                                bg='#ffffff') # 保持白色背景
            calc_lbl.grid(row=i, column=col, padx=3, pady=2, sticky='w')
            col += 1
            self.calc_labels[tab_name][name] = calc_lbl

            # Operation (使用 tk.Label 以设置隔行变色)
            # 注意：这里使用 tk.Label 替代了 ttk.Label
            op_lbl = tk.Label(frame, text=name, anchor='w', justify='left',
                              wraplength=680, bg=row_color) # 设置隔行颜色
            op_lbl.grid(row=i, column=op_col, sticky='ew', padx=6, pady=2) # sticky='ew' 使其填充 op_col
            
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def update_section_times(self, tab_name):
        mode = self.mode_var.get()
        n_cav = self.get_n_cavities()
        factor = self.get_factor()
        ops_list = {'Pre_Assembly': PRE_ASSEMBLY_OPS, 'Assembly': ASSEMBLY_OPS, 'Tuning': TUNING_OPS,
                    'Testing': TESTING_OPS, 'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS}[tab_name]

        rf_name = "RF TUNING1 estimated (if no other indication)"
        check_name = "Tuning check after cleaning"
        rf_minutes = 0.0
        if tab_name == 'Tuning':
            if self.check_vars[tab_name][rf_name].get():
                std_t_rf = self.std_time_vars[(tab_name, rf_name)].get()
                opt_t_rf = self.opt_time_dict[(tab_name, rf_name)]
                t_per_rf = std_t_rf if mode == 'STANDARD' else opt_t_rf
                qty_rf = self.quantity_vars[tab_name][rf_name].get()
                unit_rf = self.unit_dict[(tab_name, rf_name)]
                rf_minutes = (t_per_rf / 60.0 if unit_rf == 'sec' else t_per_rf) * qty_rf * n_cav
                self.std_time_vars[(tab_name, check_name)].set(0.1 * rf_minutes)
            else:
                self.std_time_vars[(tab_name, check_name)].set(self.original_std_times[(tab_name, check_name)])

        for name, _, opt_t, _, unit, mtype in ops_list:
            if not self.check_vars[tab_name][name].get():
                self.calc_labels[tab_name][name].config(text="0.00")
                continue
            std_t = self.std_time_vars[(tab_name, name)].get()
            t_per = std_t if mode == 'STANDARD' else opt_t
            qty = self.quantity_vars[tab_name][name].get()
            if tab_name == 'Tuning' and name == rf_name:
                minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty * n_cav
            else:
                minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty
            if tab_name in ['Pre_Assembly', 'Assembly']:
                parts_qty = self.parts_qty_vars[tab_name][name].get()
                minutes *= parts_qty
            self.calc_labels[tab_name][name].config(text=f"{minutes:8.2f}")

    def update_all_times(self, *args):
        for tab in self.check_vars:
            self.update_section_times(tab)

    def get_section_total(self, tab_name, ops_list):
        total = 0.0
        mode = self.mode_var.get()
        n_cav = self.get_n_cavities()
        factor = self.get_factor()
        rf_name = "RF TUNING1 estimated (if no other indication) "
        # For Tuning, ensure std_check is updated
        if tab_name == 'Tuning':
            self.update_section_times(tab_name)
        for name, _, opt_t, _, unit, mtype in ops_list:
            if not self.check_vars[tab_name][name].get(): continue
            std_t = self.std_time_vars[(tab_name, name)].get()
            t_per = std_t if mode == 'STANDARD' else opt_t
            qty = self.quantity_vars[tab_name][name].get()
            if tab_name == 'Tuning' and name == rf_name:
                base = (t_per / 60.0 if unit == 'sec' else t_per) * qty * n_cav
            else:
                base = (t_per / 60.0 if unit == 'sec' else t_per) * qty
            if tab_name in ['Pre_Assembly', 'Assembly']:
                parts_qty = self.parts_qty_vars[tab_name][name].get()
                base *= parts_qty
            if tab_name in ['Assembly', 'Tuning', 'Testing']:
                total += base * factor
            else:
                total += base
        return total

    def reset_all(self):
        self.n_cavities_var.set('0')
        self.housing_per_fg_var.set('2')
        self.mode_var.set('STANDARD')
        self.pn_var.set("")
        self.fpy_var.set('88%')
        self.custom_fpy_var.set('88%')
        self.rework_var.set('12%')
        housing = self.get_housing()
        section_ops = {'Pre_Assembly': PRE_ASSEMBLY_OPS, 'Assembly': ASSEMBLY_OPS, 'Tuning': TUNING_OPS,
                       'Testing': TESTING_OPS, 'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS}
        self.suppress_qty_modify = True
        rf_name = "RF TUNING1 estimated (if no other indication)"
        check_name = "Tuning check after cleaning"
        for tab in self.check_vars:
            for var in self.check_vars[tab].values(): var.set(False)
            self.user_modified[tab] = {}
            if tab in ['Pre_Assembly', 'Assembly']:
                for var in self.parts_qty_vars[tab].values():
                    var.set(1)
            for name, var in self.quantity_vars[tab].items():
                for op in section_ops[tab]:
                    if op[0] == name:
                        default_qty = op[3]
                        if tab == 'Tuning' and name == check_name:
                            var.set(default_qty)
                        elif tab in ['Pre_Assembly', 'Assembly'] or (tab == 'Tuning' and name == rf_name):
                            var.set(default_qty * housing)
                        else:
                            var.set(default_qty)
                        break
            for key in [k for k in self.std_time_vars if k[0] == tab]:
                self.std_time_vars[key].set(self.original_std_times[key])
        self.suppress_qty_modify = False
        self.update_all_times()
        self.results.delete(1.0, tk.END)
        self.custom_fpy_entry.config(state='disabled')

    def calculate_detailed(self):
        self.update_all_times()
        product = self.product_var.get()
        cm = self.cm_var.get()
        mode = self.mode_var.get()
        n_cav = self.get_n_cavities()
        housing = self.get_housing()
        modules = housing
        tuning_type = self.tuning_type_var.get()
        pn = self.pn_var.get() or "N/A"

        factor = self.get_factor()
        fpy = self.get_fpy()

        pre_assembly_base = self.get_section_total('Pre_Assembly', PRE_ASSEMBLY_OPS)
        pre_assembly = pre_assembly_base * factor
        assembly_base = self.get_section_total('Assembly', ASSEMBLY_OPS)
        assembly = assembly_base * factor
        tuning_base = self.get_section_total('Tuning', TUNING_OPS)
        tuning_adjusted = tuning_base * factor
        testing_base = self.get_section_total('Testing', TESTING_OPS)
        testing = testing_base * factor
        quality = self.get_section_total('Quality Check', QUALITY_OPS)
        packaging = self.get_section_total('Packaging', PACKAGING_OPS)

        rf_tuning_time_base = 0.0
        rf_tuning_time = 0.0
        if tuning_type in TUNING_TYPES and n_cav > 0:
            std_sec, opt_sec = TUNING_TYPES[tuning_type]
            rf_tuning_time_base = ((std_sec if mode == 'STANDARD' else opt_sec) / 60.0) * n_cav * modules
            rf_tuning_time = rf_tuning_time_base * factor

        total_wo = pre_assembly + assembly + rf_tuning_time + tuning_adjusted + testing + quality + packaging
        rework_rate = 1 - fpy
        rework = (total_wo - pre_assembly - quality - packaging) * rework_rate
        total = total_wo + rework

        result = f""" ROUTING CALCULATION RESULT 
{'═' * 100}
P/N            : {pn}
Product        : {product:<12} CM: {cm:<12} FPY: {self.get_fpy_display()}
Tuning Type    : {tuning_type}
Cavities × Modules : {n_cav} × {modules:<6} Housing/FG: {housing}
Mode           : {mode}

Pre_Assembly   : {pre_assembly_base:6.2f}
Assembly       : {assembly_base:6.2f}
RF Tuning      : {rf_tuning_time_base:6.2f}
Tuning         : {tuning_base:6.2f}
Testing        : {testing_base:6.2f}
Cleaning       : {0.0:6.2f}  (included in Assembly)
Quality Check  : {quality:6.2f}
Packaging      : {packaging:6.2f}
{'─' * 50}
Total w/o Rework : {total_wo:6.2f}
Rework         : {rework:6.2f}
╔{'═'*42}╗
║  FINAL ROUTING TIME : {total:7.2f} minutes      ║
╚{'═'*42}╝
"""
        self.results.delete(1.0, tk.END)
        self.results.insert(tk.END, result)

    def export_to_excel(self):
        self.update_all_times()
        try:
            pn_raw = self.pn_var.get().strip()
            pn = "".join(c if c.isalnum() or c in "._-/\\" else "_" for c in pn_raw) if pn_raw else "NoPN"
            tuning = self.tuning_type_var.get().replace(" ", "_").replace("(", "").replace(")", "")
            suggested_name = f"Routing_{pn}_{tuning}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel File", "*.xlsx")],
                title="Save Routing Result",
                initialfile=suggested_name
            )
            if not filename:
                return

            rows = []
            for tab, ops in [('Pre_Assembly', PRE_ASSEMBLY_OPS), ('Assembly', ASSEMBLY_OPS), ('Tuning', TUNING_OPS),
                            ('Testing', TESTING_OPS), ('Quality Check', QUALITY_OPS), ('Packaging', PACKAGING_OPS)]:
                for name, _, opt_t, _, unit, _ in ops:
                    if self.check_vars[tab][name].get():
                        parts_qty_val = self.parts_qty_vars[tab][name].get() if tab in self.parts_qty_vars else 1
                        rows.append({
                            "Section": tab,
                            "Operation": name,
                            "Parts Qty": parts_qty_val,
                            "Qty": self.quantity_vars[tab][name].get(),
                            "Std Time": f"{self.std_time_vars[(tab,name)].get():.2f} {unit}",
                            "Calc Time (min)": round(float(self.calc_labels[tab][name]['text']), 3)
                        })
            df_ops = pd.DataFrame(rows) if rows else pd.DataFrame({"Info": ["No operations selected"]})

            pre_assembly = self.get_section_total('Pre_Assembly', PRE_ASSEMBLY_OPS)
            assembly = self.get_section_total('Assembly', ASSEMBLY_OPS)
            tuning = self.get_section_total('Tuning', TUNING_OPS)
            testing = self.get_section_total('Testing', TESTING_OPS)
            quality = self.get_section_total('Quality Check', QUALITY_OPS)
            packaging = self.get_section_total('Packaging', PACKAGING_OPS)
            n_cav = self.get_n_cavities()
            housing = self.get_housing()
            factor = self.get_factor()
            rf_tuning_time = 0.0
            if self.tuning_type_var.get() in TUNING_TYPES and n_cav > 0:
                std_sec, opt_sec = TUNING_TYPES[self.tuning_type_var.get()]
                rf_tuning_time_base = ((std_sec if self.mode_var.get() == 'STANDARD' else opt_sec) / 60.0) * n_cav * housing
                rf_tuning_time = rf_tuning_time_base * factor

            total_wo_rework = pre_assembly + assembly + rf_tuning_time + tuning + testing + quality + packaging
            rework_rate = 1 - self.get_fpy()
            rework = (total_wo_rework - pre_assembly - quality - packaging) * rework_rate
            final_total = total_wo_rework + rework

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Configuration
                config_data = [
                    ["P/N", pn_raw or "N/A"],
                    ["Product", self.product_var.get()],
                    ["CM", self.cm_var.get()],
                    ["Mode", self.mode_var.get()],
                    ["FPY", self.get_fpy_display()],
                    ["Tuning Type", self.tuning_type_var.get()],
                    ["N° Cavities", n_cav],
                    ["Housing per FG", housing],
                    ["Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                ]
                pd.DataFrame(config_data, columns=["Parameter", "Value"]).to_excel(writer, sheet_name="Configuration", index=False)

                df_ops.to_excel(writer, sheet_name="Selected Operations", index=False)

                summary_data = [
                    ["Pre_Assembly", round(pre_assembly,3), 1.0, round(pre_assembly,3)],
                    ["Assembly", round(assembly / factor,3) if factor != 0 else 0, factor, round(assembly,3)],
                    ["RF Tuning", round(rf_tuning_time / factor,3) if factor != 0 else 0, factor, round(rf_tuning_time,3)],
                    ["Tuning", round(tuning / factor,3) if factor != 0 else 0, factor, round(tuning,3)],
                    ["Testing", round(testing / factor,3) if factor != 0 else 0, factor, round(testing,3)],
                    ["Quality Check", round(quality,3), 1.0, round(quality,3)],
                    ["Packaging", round(packaging,3), 1.0, round(packaging,3)],
                    ["Total w/o Rework", round(total_wo_rework,3), "", round(total_wo_rework,3)],
                    ["Rework", round(rework,3), "", round(rework,3)],
                    ["FINAL ROUTING TIME", round(final_total,3), "", round(final_total,3)],
                ]
                pd.DataFrame(summary_data, columns=["Item", "Base Time", "Factor", "Adjusted Time"]).to_excel(writer, sheet_name="Time Summary", index=False)

                # 美化
                wb = writer.book
                for ws in wb.worksheets:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D6EBFF", end_color="D6EBFF", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    for col in ws.columns:
                        max_len = 0
                        column = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[column].width = min(max_len + 4, 60)
                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.border = border

            messagebox.showinfo("Success", f"Excel has been successfully exported!\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    set_app_icon(root)
    app = RoutingApp(root)
    root.mainloop()