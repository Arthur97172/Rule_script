# -*- coding: utf-8 -*-
"""
Routing Estimation Tool - 工业级终极完整版
功能：
1. 新增 P/N 输入框
2. 导出文件名自动包含 P/N + Tuning Type
3. 一键导出美观 Excel（3个Sheet）
4. 完整工序数据 + 实时计算 + Reset All
5. 滚轮同步滚动（Windows）
6. Assembly/Testing 等表格隔行变色（白/浅灰）
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import scrolledtext
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================================= 数据定义 =================================
REWORK_FIXED = 12.666666666666668

PRODUCTS = ['TMA', 'DPX', 'TPX', 'PPX', 'QPX', 'IMF', 'LLC']

DELTAS = {
    'TIGER': {p: [0.0, 0.0, 0.0] for p in PRODUCTS},
    'TATFOOK': {
        'TMA': [0.39, 0.24, 0.10], 'DPX': [0.39, 0.35, 0.28], 'TPX': [0.43, 0.68, -0.15],
        'PPX': [0.04, 0.18, -0.05], 'QPX': [-0.08, -0.16, -0.17],
        'IMF': [0.02, 0.04, -0.31], 'LLC': [-0.16, -0.13, 0.0]
    },
    'LUXSHARE': {
        'TMA': [1.39, 1.24, 1.1], 'DPX': [1.39, 1.35, 1.28], 'TPX': [1.43, 1.68, 0.85],
        'PPX': [1.04, 1.18, 0.95], 'QPX': [0.92, 0.84, 0.83],
        'IMF': [1.02, 1.04, 0.69], 'LLC': [0.84, 0.87, 0.0]
    }
}

FPY = {'DPX': 0.9, 'TPX': 0.88, 'QPX': 0.85, 'TMA': 0.9, 'IMF': 0.85, 'LLC': 0.85, 'PPX': 0.85}

ASSEMBLY_OPS = [
    ("tuning screw glued to tuners or plug", 12, 12, 2, 'sec', 'housing'),
    ("lid cover+tuning screw+nut+washer", 8, 8, 2, 'sec', 'housing'),
    ("connector+conductor soldered", 45, 45, 2, 'sec', 'housing'),
    ("connector+conductor screwed", 5, 5, 2, 'sec', 'housing'),
    ("low pass(or bias tee or coils) + thermalshrinking", 120, 120, 2, 'sec', 'housing'),
    ("bias coil group", 30, 30, 2, 'sec', 'housing'),
    ("lid cover + pin", 30, 30, 2, 'sec', 'housing'),
    ("lid cover + line screwed", 5, 5, 2, 'sec', 'housing'),
    ("lid cover + line soldered", 67, 67, 2, 'sec', 'housing'),
    ("lid cover + connector", 15, 15, 2, 'sec', 'housing'),
    ("lid cover + connector press in", 60, 60, 2, 'sec', 'housing'),
    ("label for s/n", 5, 5, 2, 'sec', 'housing'),
    ("housing + pin", 10, 10, 2, 'sec', 'housing'),
    ("housing + connector flanged", 45, 45, 2, 'sec', 'housing'),
    ("housing + connector screwed", 30, 30, 2, 'sec', 'housing'),
    ("housing + junction screwed (time x 2 screws)", 10, 10, 2, 'sec', 'housing'),
    ("housing + junction soldered by SOLDER STATION", 30, 30, 2, 'sec', 'housing'),
    ("housing + junction soldered INDUCTION", 55, 55, 2, 'sec', 'housing'),
    ("housing + line soldered by SOLDER STATION (complex)", 150, 150, 2, 'sec', 'housing'),
    ("housing + line soldered INDUCTION (complex)", 150, 150, 2, 'sec', 'housing'),
    ("housing + low pass(or bias tee)", 10, 10, 2, 'sec', 'housing'),
    ("ret connector", 20, 20, 2, 'sec', 'housing'),
    ("resonator", 7, 7, 2, 'sec', 'housing'),
    ("housing + probe (including support fixing)", 9, 9, 2, 'sec', 'housing'),
    ("apply solder paste for LID SOLDERING process", 300, 300, 2, 'sec', 'housing'),
    ("tuning lid cover screwing", 3, 3, 2, 'sec', 'housing'),
    ("PWBA installation (including soldering)", 8, 8, 2, 'sec', 'housing'),
    ("Cable sub-assembly(or LED)", 15, 15, 1, 'sec', 'housing'),
    ("Outer cover (time x screw)", 4, 4, 2, 'sec', 'housing'),
    ("PWBA+Nickel Silver Cover placing", 3, 3, 2, 'sec', 'housing'),
    ("IP outer label ASSEMBLY", 60, 60, 2, 'sec', 'housing'),
    ("housing + venting membrane", 10, 10, 1, 'sec', 'housing'),
    ("housing + bracket +screw + nut + washer", 120, 120, 1, 'sec', 'housing'),
]

TESTING_OPS = [
    ("PCBA programming (FW uploading)", 0.0, 0.0, 1, 'min', 'none'),
    ("TUNING 2 (E35)", 8.0, 8.0, 1, 'min', 'none'),
    ("NOISE FIGURE(mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("IP3(optional for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("AC/DC or Generic Measures", 3.0, 3.0, 1, 'min', 'none'),
    ("IMD (STANDARD)", 1.5, 0.0, 1, 'min', 'none'),
    ("IMD (OPTIMIZED)", 0.0, 1.5, 1, 'min', 'none'),
    ("LINEARIZATION (mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("PID(mandatory for TMA)", 3.0, 3.0, 1, 'min', 'none'),
    ("TDR on TMS (FINAL TEST)", 3.0, 3.0, 1, 'min', 'none'),
    ("PRESSURE TEST", 2.0, 2.0, 1, 'min', 'none'),
    ("LED", 2.0, 2.0, 1, 'min', 'none'),
]

CLEANING_OPS = [
    ("tuning lid open and close (time x screw)", 30, 30, 1, 'sec', 'housing'),
    ("Cleaning", 4.0, 4.0, 2, 'min', 'none'),
    ("tuning check after cleaning", 5.0, 5.0, 1, 'min', 'none'),
]

QUALITY_OPS = [("Final check 100% (VISUAL INSPECTION)", 1.0, 1.0, 2, 'min', 'none')]
PACKAGING_OPS = [("Packaging", 5.2, 5.2, 1, 'min', 'none')]

TUNING_TYPES = {
    'multiplexer': (80, 72),
    'MPX (from PPX on)': (100, 90),
    'TMA': (100, 90),
    'LLC': (120, 108),
    'IMF': (120, 108),
    'demanding perf': (120, 108),
}

# ==============================================================================

class RoutingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Routing Estimation Tool - Industrial Ultimate")
        self.root.geometry("1080x900")
        self.root.minsize(900, 720)

        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill='both', expand=True, padx=15, pady=12)

        # 新增 P/N
        self.pn_var = tk.StringVar(value="")

        self.product_var = tk.StringVar(value='TPX')
        self.cm_var = tk.StringVar(value='TIGER')
        self.mode_var = tk.StringVar(value='STANDARD')
        self.fpy_var = tk.StringVar(value='0.88')
        self.n_cavities_var = tk.IntVar(value=0)
        self.housing_per_fg_var = tk.IntVar(value=2)
        self.global_modules_var = tk.IntVar(value=2)
        self.tuning_type_var = tk.StringVar(value='multiplexer')

        self.original_std_times = {}
        self.check_vars = {}
        self.quantity_vars = {}
        self.std_time_vars = {}
        self.opt_time_dict = {}
        self.unit_dict = {}
        self.multi_type_dict = {}
        self.calc_labels = {}

        self.setup_data()
        self.setup_ui()

        self.mode_var.trace('w', self.update_all_times)
        self.housing_per_fg_var.trace('w', self.update_all_times)
        self.product_var.trace('w', self.on_product_change)

        self.update_all_times()
        self.on_product_change()

    ###############################################################
    # 数据初始化
    ###############################################################
    def setup_data(self):
        sections = {
            'Assembly': ASSEMBLY_OPS, 'Testing': TESTING_OPS,
            'Cleaning': CLEANING_OPS, 'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS
        }
        for tab, ops in sections.items():
            self.check_vars[tab] = {}
            self.quantity_vars[tab] = {}
            for op in ops:
                name, std_t, opt_t, qty, unit, mtype = op
                self.check_vars[tab][name] = tk.BooleanVar(value=False)
                self.quantity_vars[tab][name] = tk.IntVar(value=qty)
                self.std_time_vars[(tab, name)] = tk.DoubleVar(value=std_t)
                self.original_std_times[(tab, name)] = std_t
                self.opt_time_dict[(tab, name)] = opt_t
                self.unit_dict[(tab, name)] = unit
                self.multi_type_dict[(tab, name)] = mtype

                for var in [self.check_vars[tab][name], self.quantity_vars[tab][name], self.std_time_vars[(tab, name)]]:
                    var.trace('w', lambda *_, t=tab: self.update_section_times(t))

    ###############################################################
    # 根据产品改变 FPY
    ###############################################################
    def on_product_change(self, *args):
        prod = self.product_var.get()
        if prod in FPY:
            self.fpy_var.set(f"{FPY[prod]:.2f}")

    ###############################################################
    # UI 构建
    ###############################################################
    def setup_ui(self):

        #========== Basic Configuration ==========
        basic_frame = ttk.LabelFrame(self.main_frame, text=" Basic Configuration ", padding=12)
        basic_frame.pack(fill='x', pady=(0, 8))

        # 第一行：Product / CM / Tuning Type
        ttk.Label(basic_frame, text="Product:").grid(row=0, column=0, sticky='e', padx=(0,8))
        ttk.Combobox(basic_frame, textvariable=self.product_var, values=PRODUCTS, width=15).grid(row=0, column=1, sticky='w', padx=(0,30))

        ttk.Label(basic_frame, text="CM:").grid(row=0, column=2, sticky='e', padx=(20,8))
        ttk.Combobox(basic_frame, textvariable=self.cm_var, values=list(DELTAS.keys()), width=15).grid(row=0, column=3, sticky='w', padx=(0,30))

        ttk.Label(basic_frame, text="Tuning Type:").grid(row=0, column=4, sticky='e', padx=(20,8))
        ttk.Combobox(basic_frame, textvariable=self.tuning_type_var, values=list(TUNING_TYPES.keys()), width=26).grid(row=0, column=5, sticky='w')

        # 第二行：Mode / FPY
        ttk.Label(basic_frame, text="Mode:").grid(row=1, column=0, sticky='e', padx=(0,8), pady=(10,0))
        ttk.Combobox(basic_frame, textvariable=self.mode_var, values=['STANDARD', 'OPTIMIZED'], width=15).grid(row=1, column=1, sticky='w', padx=(0,30), pady=(10,0))

        ttk.Label(basic_frame, text="FPY:").grid(row=1, column=2, sticky='e', padx=(20,8), pady=(10,0))
        ttk.Combobox(basic_frame, textvariable=self.fpy_var, values=[f"{v:.2f}" for v in sorted(set(FPY.values()))], width=15).grid(row=1, column=3, sticky='w', pady=(10,0))

        # P/N
        ttk.Label(basic_frame, text="P/N:", font=('Arial', 10, 'bold')).grid(row=1, column=4, sticky='e', padx=(50,8), pady=(10,0))
        ttk.Entry(basic_frame, textvariable=self.pn_var, width=28, font=('Arial', 10)).grid(row=1, column=5, sticky='w', pady=(10,0))

        #========== Routing Parameters ==========
        param_frame = ttk.LabelFrame(self.main_frame, text=" Routing Parameters ", padding=12)
        param_frame.pack(fill='x', pady=(0, 12))

        ttk.Label(param_frame, text="N° Cavities:").grid(row=0, column=0, sticky='w', padx=(0,10))
        ttk.Entry(param_frame, textvariable=self.n_cavities_var, width=12, justify='center').grid(row=0, column=1, padx=(0,40))

        ttk.Label(param_frame, text="Housing per FG:").grid(row=0, column=2, sticky='w', padx=(20,10))
        ttk.Entry(param_frame, textvariable=self.housing_per_fg_var, width=12, justify='center').grid(row=0, column=3, padx=(0,40))

        ttk.Label(param_frame, text="Modules per FG (tuning):").grid(row=0, column=4, sticky='w', padx=(20,10))
        ttk.Entry(param_frame, textvariable=self.global_modules_var, width=12, justify='center').grid(row=0, column=5)

        #========== Tabs ==========
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill='both', expand=True, pady=(0,10))

        for tab_name, ops in [('Assembly', ASSEMBLY_OPS), ('Testing', TESTING_OPS),
                              ('Cleaning', CLEANING_OPS), ('Quality Check', QUALITY_OPS),
                              ('Packaging', PACKAGING_OPS)]:
            self.create_tab(tab_name, ops)

        #========== Buttons ==========
        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.pack(pady=12)
        ttk.Button(btn_frame, text="Calculate Total Time", command=self.calculate_detailed, width=24).pack(side='left', padx=12)
        ttk.Button(btn_frame, text="Reset All", command=self.reset_all, width=20).pack(side='left', padx=12)
        ttk.Button(btn_frame, text="Export to Excel", command=self.export_to_excel, style="Accent.TButton", width=24).pack(side='left', padx=12)

        #========== Result Display ==========
        self.results = scrolledtext.ScrolledText(self.main_frame, height=16, font=("Courier New", 10), bg='#f8f8f8')
        self.results.pack(fill='both', expand=True, padx=10, pady=(0,12))

        # Windows 滚轮绑定
        self.root.bind_all("<MouseWheel>", self.on_mousewheel)

    ###############################################################
    # Windows 滚轮事件处理（同步滚动）
    ###############################################################
    def on_mousewheel(self, event):
        delta = int(-1 * (event.delta / 120))

        if self.results.winfo_exists():
            self.results.yview_scroll(delta, "units")

        current_tab = self.notebook.nametowidget(self.notebook.select())
        for child in current_tab.winfo_children():
            if isinstance(child, tk.Canvas):
                child.yview_scroll(delta, "units")

    ###############################################################
    # 创建 Tab（含隔行变色）
    ###############################################################
    def create_tab(self, tab_name, ops_list):
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text=f" {tab_name} ")
        self.calc_labels[tab_name] = {}

        canvas = tk.Canvas(tab)
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        frame = ttk.Frame(canvas)
        frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Header
        headers = ["Select", "Qty", "Std Time", "Calc Time (min)", "Operation"]
        for col, text in enumerate(headers):
            lbl = tk.Label(frame, text=text, font=('Courier New', 9, 'bold'),
                           bg='#d9d9d9', anchor='w', padx=4)
            lbl.grid(row=0, column=col, padx=3, pady=5, sticky='w')

        # Operation 列占据剩余空间
        frame.grid_columnconfigure(4, weight=1)

        # 数据行
        for i, (name, std_t, opt_t, qty, unit, mtype) in enumerate(ops_list, 1):

            # 仅 Operation 列隔行变色
            row_color = "#ffffff" if i % 2 == 0 else "#f0f2f5"
            key = (tab_name, name)

            # Select（保持默认 ttk 外观）
            ttk.Checkbutton(frame, variable=self.check_vars[tab_name][name]).grid(
                row=i, column=0, padx=6, pady=2, sticky='w')

            # Qty
            ttk.Entry(frame, textvariable=self.quantity_vars[tab_name][name],
                      width=6, justify='center').grid(row=i, column=1, padx=3, pady=2, sticky='w')

            # Std Time + unit
            tf = ttk.Frame(frame)
            tf.grid(row=i, column=2, padx=3, pady=2, sticky='w')
            ttk.Entry(tf, textvariable=self.std_time_vars[key],
                      width=8, justify='center').pack(side='left')
            ttk.Label(tf, text=unit, width=4).pack(side='left')

            # Calc time
            calc_lbl = tk.Label(frame, text="0.00",
                    font=('Courier New', 10),
                    fg='darkgreen',   # 恢复绿色
                    bg='white')       # 保持整齐
            calc_lbl.grid(row=i, column=3, padx=3, pady=2, sticky='w')
            self.calc_labels[tab_name][name] = calc_lbl

            # Operation（唯一有隔行颜色的列）
            op_lbl = tk.Label(frame, text=name, anchor='w', justify='left',
                              width=70, wraplength=680, bg=row_color)
            op_lbl.grid(row=i, column=4, sticky='w', padx=6, pady=2)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")



    def update_section_times(self, tab_name):
        mode = self.mode_var.get()
        housing = self.housing_per_fg_var.get()
        ops_list = {'Assembly': ASSEMBLY_OPS, 'Testing': TESTING_OPS, 'Cleaning': CLEANING_OPS,
                    'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS}[tab_name]

        for name, _, opt_t, _, unit, mtype in ops_list:
            if not self.check_vars[tab_name][name].get():
                self.calc_labels[tab_name][name].config(text="0.00")
                continue
            std_t = self.std_time_vars[(tab_name, name)].get()
            t_per = std_t if mode == 'STANDARD' else opt_t
            qty = self.quantity_vars[tab_name][name].get()
            multi = housing if mtype == 'housing' else 1
            minutes = (t_per / 60.0 if unit == 'sec' else t_per) * qty * multi
            self.calc_labels[tab_name][name].config(text=f"{minutes:8.2f}")

    def update_all_times(self, *args):
        for tab in self.check_vars:
            self.update_section_times(tab)

    def get_section_total(self, tab_name, ops_list):
        total = 0.0
        mode = self.mode_var.get()
        housing = self.housing_per_fg_var.get()
        for name, _, opt_t, _, unit, mtype in ops_list:
            if not self.check_vars[tab_name][name].get(): continue
            std_t = self.std_time_vars[(tab_name, name)].get()
            t_per = std_t if mode == 'STANDARD' else opt_t
            qty = self.quantity_vars[tab_name][name].get()
            multi = housing if mtype == 'housing' else 1
            total += (t_per / 60.0 if unit == 'sec' else t_per) * qty * multi
        return total

    def reset_all(self):
        for tab in self.check_vars:
            for var in self.check_vars[tab].values(): var.set(False)
            for name, var in self.quantity_vars[tab].items():
                for op in {'Assembly': ASSEMBLY_OPS, 'Testing': TESTING_OPS, 'Cleaning': CLEANING_OPS,
                           'Quality Check': QUALITY_OPS, 'Packaging': PACKAGING_OPS}[tab]:
                    if op[0] == name:
                        var.set(op[3])
                        break
            for key in [k for k in self.std_time_vars if k[0] == tab]:
                self.std_time_vars[key].set(self.original_std_times[key])
        self.n_cavities_var.set(0)
        self.housing_per_fg_var.set(2)
        self.global_modules_var.set(2)
        self.mode_var.set('STANDARD')
        self.pn_var.set("")
        self.update_all_times()
        self.results.delete(1.0, tk.END)

    def calculate_detailed(self):
        product = self.product_var.get()
        cm = self.cm_var.get()
        mode = self.mode_var.get()
        n_cav = self.n_cavities_var.get()
        housing = self.housing_per_fg_var.get()
        modules = self.global_modules_var.get()
        tuning_type = self.tuning_type_var.get()
        pn = self.pn_var.get() or "N/A"

        deltas = DELTAS.get(cm, DELTAS['TIGER'])[product]

        assembly = self.get_section_total('Assembly', ASSEMBLY_OPS)
        testing = self.get_section_total('Testing', TESTING_OPS)
        cleaning = self.get_section_total('Cleaning', CLEANING_OPS)
        quality = self.get_section_total('Quality Check', QUALITY_OPS)
        packaging = self.get_section_total('Packaging', PACKAGING_OPS)

        tuning_time = 0.0
        if tuning_type in TUNING_TYPES and n_cav > 0:
            std_sec, opt_sec = TUNING_TYPES[tuning_type]
            tuning_time = ((std_sec if mode == 'STANDARD' else opt_sec) / 60.0) * n_cav * modules

        total_wo = assembly + cleaning + tuning_time + testing + quality + packaging
        total = total_wo + REWORK_FIXED

        result = f""" ROUTING CALCULATION RESULT 
{'═' * 100}
P/N            : {pn}
Product        : {product:<12} CM: {cm:<12} FPY: {self.fpy_var.get()}
Tuning Type    : {tuning_type}
Cavities × Modules : {n_cav} × {modules:<6} Housing/FG: {housing}
Mode           : {mode}

Assembly       : {assembly:6.2f}  + Δ{deltas[0]:+.2f} → {assembly + deltas[0]:6.2f}
RF Tuning      : {tuning_time:6.2f}  + Δ{deltas[1]:+.2f} → {tuning_time + deltas[1]:6.2f}
Testing        : {testing:6.2f}  + Δ{deltas[2]:+.2f} → {testing + deltas[2]:6.2f}
Cleaning       : {cleaning:6.2f}
Quality Check  : {quality:6.2f}
Packaging      : {packaging:6.2f}
{'─' * 50}
Total w/o Rework : {total_wo:6.2f}
Rework (fixed)   : {REWORK_FIXED:6.2f}
╔{'═'*42}╗
║  FINAL ROUTING TIME : {total:7.2f} minutes      ║
╚{'═'*42}╝
"""
        self.results.delete(1.0, tk.END)
        self.results.insert(tk.END, result)

    def export_to_excel(self):
        try:
            pn_raw = self.pn_var.get().strip()
            pn = "".join(c if c.isalnum() or c in "._-/\\" else "_" for c in pn_raw) if pn_raw else "NoPN"
            tuning = self.tuning_type_var.get().replace(" ", "_").replace("(", "").replace(")", "")
            suggested_name = f"Routing_{pn}_{tuning}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                title="保存 Routing 结果",
                initialfile=suggested_name
            )
            if not filename:
                return

            # 收集勾选工序
            rows = []
            for tab, ops in [('Assembly', ASSEMBLY_OPS), ('Testing', TESTING_OPS),
                            ('Cleaning', CLEANING_OPS), ('Quality Check', QUALITY_OPS),
                            ('Packaging', PACKAGING_OPS)]:
                for name, _, opt_t, _, unit, _ in ops:
                    if self.check_vars[tab][name].get():
                        rows.append({
                            "Section": tab,
                            "Operation": name,
                            "Qty": self.quantity_vars[tab][name].get(),
                            "Std Time": f"{self.std_time_vars[(tab,name)].get():.2f} {unit}",
                            "Calc Time (min)": round(float(self.calc_labels[tab][name]['text']), 3)
                        })
            df_ops = pd.DataFrame(rows) if rows else pd.DataFrame({"Info": ["No operations selected"]})

            # 计算汇总
            assembly = self.get_section_total('Assembly', ASSEMBLY_OPS)
            testing = self.get_section_total('Testing', TESTING_OPS)
            cleaning = self.get_section_total('Cleaning', CLEANING_OPS)
            quality = self.get_section_total('Quality Check', QUALITY_OPS)
            packaging = self.get_section_total('Packaging', PACKAGING_OPS)
            tuning_time = 0.0
            if self.tuning_type_var.get() in TUNING_TYPES and self.n_cavities_var.get() > 0:
                std_sec, opt_sec = TUNING_TYPES[self.tuning_type_var.get()]
                tuning_time = ((std_sec if self.mode_var.get() == 'STANDARD' else opt_sec) / 60.0) * \
                              self.n_cavities_var.get() * self.global_modules_var.get()

            deltas = DELTAS.get(self.cm_var.get(), DELTAS['TIGER'])[self.product_var.get()]
            total_wo_rework = assembly + cleaning + tuning_time + testing + quality + packaging
            final_total = total_wo_rework + REWORK_FIXED

            # 写入 Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Configuration
                config_data = [
                    ["P/N", pn_raw or "N/A"],
                    ["Product", self.product_var.get()],
                    ["CM", self.cm_var.get()],
                    ["Mode", self.mode_var.get()],
                    ["FPY", self.fpy_var.get()],
                    ["Tuning Type", self.tuning_type_var.get()],
                    ["N° Cavities", self.n_cavities_var.get()],
                    ["Housing per FG", self.housing_per_fg_var.get()],
                    ["Modules per FG", self.global_modules_var.get()],
                    ["Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                ]
                pd.DataFrame(config_data, columns=["Parameter", "Value"]).to_excel(writer, sheet_name="Configuration", index=False)

                df_ops.to_excel(writer, sheet_name="Selected Operations", index=False)

                summary_data = [
                    ["Assembly", round(assembly,3), deltas[0], round(assembly + deltas[0],3)],
                    ["RF Tuning", round(tuning_time,3), deltas[1], round(tuning_time + deltas[1],3)],
                    ["Testing", round(testing,3), deltas[2], round(testing + deltas[2],3)],
                    ["Cleaning", round(cleaning,3), 0, round(cleaning,3)],
                    ["Quality Check", round(quality,3), 0, round(quality,3)],
                    ["Packaging", round(packaging,3), 0, round(packaging,3)],
                    ["Total w/o Rework", round(total_wo_rework,3), "", round(total_wo_rework,3)],
                    ["Rework (fixed)", REWORK_FIXED, "", REWORK_FIXED],
                    ["FINAL ROUTING TIME", round(final_total,3), "", round(final_total,3)],
                ]
                pd.DataFrame(summary_data, columns=["Item", "Base Time", "Delta", "Final Time"]).to_excel(writer, sheet_name="Time Summary", index=False)

                # 美化
                wb = writer.book
                for ws in wb.worksheets:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D6EBFF", end_color="D6EBFF", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    for col in ws.columns:
                        max_len = 0
                        column = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[column].width = min(max_len + 4, 60)
                    thin = Side(border_style="thin")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.border = border

            messagebox.showinfo("成功", f"Excel 已成功导出！\n{filename}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = RoutingApp(root)
    root.mainloop()